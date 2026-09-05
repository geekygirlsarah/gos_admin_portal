from decimal import Decimal

from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import Http404
from django.shortcuts import get_object_or_404, redirect
from django.urls import reverse
from django.utils import timezone
from django.views.generic import CreateView, DeleteView, ListView, UpdateView, View

from orders.forms import OrderForm
from orders.models import PurchaseOrder
from programs.models import Program
from programs.permission_views import (
    LeadMentorRequiredMixin,
    can_user_delete,
    get_user_role,
    user_is_mentor,
)
from programs.views.mixins import (
    DynamicReadPermissionMixin,
    DynamicWritePermissionMixin,
)


class OrderProgramMixin:
    """Resolves ``self.program`` and gates access on the program's feature
    toggle.

    Students/Parents only reach the orders pages when the program has the
    ``orders`` feature enabled; mentors and Lead Mentors may always access it
    (the orders list is org-wide).
    """

    def dispatch(self, request, *args, **kwargs):
        self.program = get_object_or_404(Program, pk=kwargs.get("program_id"))
        role = get_user_role(request.user)
        is_mentor = user_is_mentor(request.user) or role == "LeadMentor"
        if not self.program.features.filter(key="orders").exists() and not is_mentor:
            raise Http404("Order requests are not enabled for this program.")
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["program"] = self.program
        context["user_role"] = get_user_role(self.request.user)
        return context

    def get_success_url(self):
        return reverse("orders:order_list", kwargs={"program_id": self.program.id})


def _group_orders(orders):
    """Group a queryset of orders by program, preserving first-seen order.

    Returns a list of ``{"program": Program|None, "orders": [...], "total": Decimal|None}``.
    """
    groups = []
    index = {}
    for order in orders:
        key = order.program_id or 0
        group = index.get(key)
        if group is None:
            group = {"program": order.program, "orders": [], "total": None}
            index[key] = group
            groups.append(group)
        group["orders"].append(order)
    for group in groups:
        totals = [o.total for o in group["orders"] if o.total is not None]
        if totals:
            group["total"] = sum(totals, Decimal("0"))
    return groups


def _export_orders_csv(orders):
    import csv

    from django.http import HttpResponse

    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="order_requests.csv"'
    writer = csv.writer(response)
    writer.writerow(
        [
            "Program",
            "Item",
            "Quantity",
            "Unit Price",
            "Total",
            "Link",
            "Requested By",
            "Requested On",
            "Notes",
            "Status",
            "Ordered On",
        ]
    )
    for order in orders.iterator(chunk_size=500):
        writer.writerow(
            [
                order.program.name if order.program else "",
                order.item_name,
                order.quantity_normalized,
                order.unit_price if order.unit_price is not None else "",
                order.total if order.total is not None else "",
                order.url,
                order.requested_by_name,
                (
                    timezone.localtime(order.created_at).strftime("%Y-%m-%d %H:%M")
                    if order.created_at
                    else ""
                ),
                order.notes,
                order.get_status_display(),
                (
                    timezone.localtime(order.ordered_at).strftime("%Y-%m-%d %H:%M")
                    if order.ordered_at
                    else ""
                ),
            ]
        )
    return response


def _export_orders_xlsx(orders):
    from io import BytesIO

    from django.http import HttpResponse
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Order Requests"
    ws.append(
        [
            "Program",
            "Item",
            "Quantity",
            "Unit Price",
            "Total",
            "Link",
            "Requested By",
            "Requested On",
            "Notes",
            "Status",
            "Ordered On",
        ]
    )
    for order in orders.iterator(chunk_size=500):
        ws.append(
            [
                order.program.name if order.program else "",
                order.item_name,
                float(order.quantity) if order.quantity is not None else "",
                float(order.unit_price) if order.unit_price is not None else "",
                float(order.total) if order.total is not None else "",
                order.url,
                order.requested_by_name,
                (
                    timezone.localtime(order.created_at).strftime("%Y-%m-%d %H:%M")
                    if order.created_at
                    else ""
                ),
                order.notes,
                order.get_status_display(),
                (
                    timezone.localtime(order.ordered_at).strftime("%Y-%m-%d %H:%M")
                    if order.ordered_at
                    else ""
                ),
            ]
        )
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=11):
        for cell in row:
            if cell.column in (2, 6, 9):
                cell.alignment = cell.alignment.copy(vertical="top")

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="order_requests.xlsx"'
    return response


class _OrderExportMixin:
    """Handles ``?export=csv`` / ``?export=xlsx`` downloads (Lead Mentors only)."""

    def _handle_export(self, orders):
        export = self.request.GET.get("export")
        if not export:
            return None
        if get_user_role(self.request.user) != "LeadMentor":
            messages.error(self.request, "Only Lead Mentors can export orders.")
            return redirect("orders:order_list", program_id=self.program.id)
        if export == "csv":
            return _export_orders_csv(orders)
        if export == "xlsx":
            return _export_orders_xlsx(orders)
        return None

    def get(self, request, *args, **kwargs):
        orders = self.get_queryset()
        response = self._handle_export(orders)
        if response is not None:
            return response
        return super().get(request, *args, **kwargs)


class OrderListView(
    LoginRequiredMixin,
    OrderProgramMixin,
    _OrderExportMixin,
    DynamicReadPermissionMixin,
    ListView,
):
    model = PurchaseOrder
    template_name = "orders/order_list.html"
    context_object_name = "orders"
    section = "orders"

    def get_queryset(self):
        return (
            PurchaseOrder.objects.filter(status=PurchaseOrder.STATUS_PENDING)
            .select_related("program", "created_by", "ordered_by")
            .order_by("-created_at")
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["order_groups"] = _group_orders(context["orders"])
        context["is_lead"] = context["user_role"] == "LeadMentor"
        context["page_title"] = "Order Requests"
        return context


class OrderArchiveView(
    LoginRequiredMixin,
    OrderProgramMixin,
    _OrderExportMixin,
    DynamicReadPermissionMixin,
    ListView,
):
    """Ordered (archived) orders, kept for budget/finance purposes."""

    model = PurchaseOrder
    template_name = "orders/order_archive.html"
    context_object_name = "orders"
    section = "orders"

    def get_queryset(self):
        return (
            PurchaseOrder.objects.filter(status=PurchaseOrder.STATUS_ORDERED)
            .select_related("program", "created_by", "ordered_by")
            .order_by("-ordered_at", "-created_at")
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["order_groups"] = _group_orders(context["orders"])
        context["is_lead"] = context["user_role"] == "LeadMentor"
        context["page_title"] = "Order Archive"
        return context


class OrderCreateView(
    LoginRequiredMixin, OrderProgramMixin, DynamicWritePermissionMixin, CreateView
):
    model = PurchaseOrder
    form_class = OrderForm
    template_name = "orders/order_form.html"
    section = "orders"

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        form.instance.program = self.program
        messages.success(self.request, "Order submitted. A Lead Mentor will review it.")
        return super().form_valid(form)


class OrderUpdateView(
    LoginRequiredMixin, OrderProgramMixin, DynamicWritePermissionMixin, UpdateView
):
    model = PurchaseOrder
    form_class = OrderForm
    template_name = "orders/order_form.html"
    section = "orders"

    def get_queryset(self):
        # Org-wide queryset; creator-only editing is enforced by
        # ``can_user_write('orders', obj)`` for non-Lead-Mentors.
        return PurchaseOrder.objects.select_related("program")

    def form_valid(self, form):
        messages.success(self.request, "Order updated.")
        return super().form_valid(form)


class OrderDeleteView(LoginRequiredMixin, OrderProgramMixin, DeleteView):
    model = PurchaseOrder
    template_name = "orders/order_confirm_delete.html"

    def get_queryset(self):
        return PurchaseOrder.objects.all()

    def dispatch(self, request, *args, **kwargs):
        if not can_user_delete(request.user, "orders", self.get_object()):
            messages.error(request, "You do not have permission to delete that order.")
            return redirect("orders:order_list", program_id=kwargs.get("program_id"))
        return super().dispatch(request, *args, **kwargs)


class OrderMarkOrderedView(
    LoginRequiredMixin, OrderProgramMixin, LeadMentorRequiredMixin, View
):
    """Lead Mentor marks a pending request as ordered (moves it to the archive)."""

    def post(self, request, program_id, pk):
        order = get_object_or_404(PurchaseOrder, pk=pk)
        if order.status == PurchaseOrder.STATUS_PENDING:
            order.status = PurchaseOrder.STATUS_ORDERED
            order.ordered_at = timezone.now()
            order.ordered_by = request.user
            order.save(update_fields=["status", "ordered_at", "ordered_by"])
            messages.success(
                request,
                f"'{order.item_name}' marked as ordered and moved to the archive.",
            )
        else:
            messages.info(request, "That order has already been marked as ordered.")
        return redirect("orders:order_list", program_id=self.program.id)


class OrderMarkPendingView(
    LoginRequiredMixin, OrderProgramMixin, LeadMentorRequiredMixin, View
):
    """Reopens an archived order (undoes 'mark ordered')."""

    def post(self, request, program_id, pk):
        order = get_object_or_404(PurchaseOrder, pk=pk)
        if order.status == PurchaseOrder.STATUS_ORDERED:
            order.status = PurchaseOrder.STATUS_PENDING
            order.ordered_at = None
            order.ordered_by = None
            order.save(update_fields=["status", "ordered_at", "ordered_by"])
            messages.success(
                request,
                f"'{order.item_name}' moved back to the pending list.",
            )
        else:
            messages.info(request, "That order is not archived.")
        return redirect("orders:order_archive", program_id=self.program.id)
