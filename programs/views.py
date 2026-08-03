import datetime
import logging
import mimetypes
import os
from decimal import ROUND_HALF_DOWN, Decimal, InvalidOperation

import cssutils
from django.conf import settings
from django.contrib import messages
from django.contrib.auth.mixins import (
    LoginRequiredMixin,
    PermissionRequiredMixin,
    UserPassesTestMixin,
)
from django.core.mail import EmailMultiAlternatives, get_connection
from django.db.models import Value
from django.db.models.functions import Coalesce, Lower, NullIf
from django.http import FileResponse, Http404, HttpResponseRedirect, QueryDict
from django.shortcuts import get_object_or_404, redirect, render
from django.template.loader import render_to_string
from django.urls import reverse
from django.utils import timezone
from django.utils.html import strip_tags
from django.views.generic import (
    CreateView,
    DetailView,
    ListView,
    UpdateView,
    View,
)
from premailer import transform

from audit.events import AuditEvent
from audit.mixins import SensitiveDataViewMixin
from audit.service import log_event
from programs.constants import RELATIONSHIP_CHOICES

from .forms import (
    AddExistingStudentToProgramForm,
    AdultForm,
    FeeAssignmentEditForm,
    FeeForm,
    PaymentForm,
    ProgramDocumentForm,
    ProgramEmailBalancesForm,
    ProgramEmailForm,
    ProgramForm,
    QuickCreateStudentForm,
    SchoolForm,
    SlidingScaleApplicationForm,
    SlidingScaleForm,
    StudentForm,
)
from .models import (
    Adult,
    AdultStudentRelationship,
    Crew,
    Enrollment,
    Fee,
    Payment,
    Program,
    ProgramDocument,
    RaceEthnicity,
    School,
    SlidingScale,
    SlidingScaleSettings,
    Student,
    SubTeam,
    TaxForm,
    Team,
)
from .permission_views import (
    LeadMentorRequiredMixin,
    PassUserToFormMixin,
    can_user_read,
    can_user_write,
    get_user_role,
)
from .utils import (
    compute_sliding_discount_rounded,
    get_safe_url,
    get_student_balance_data,
    get_student_program_balance,
    redirect_back,
)

cssutils.log.setLevel(logging.WARNING)

logger = logging.getLogger("programs.email")
forms_logger = logging.getLogger("programs.forms")


class DynamicPermissionMixin(UserPassesTestMixin):
    section = None
    permission_type = "read"  # 'read' or 'write'

    def test_func(self):
        if not self.section:
            return True
        obj = getattr(self, "object", None)
        # Avoid calling get_object on CreateViews where 'pk' in URL refers to a parent
        if not obj and hasattr(self, "get_object") and not isinstance(self, CreateView):
            try:
                obj = self.get_object()
            except Http404:
                # If object doesn't exist, we treat it as a permission failure
                # so it redirects to dashboard with an error instead of 404ing
                return False
            except Exception:  # nosec B110
                pass
        if self.permission_type == "write":
            return can_user_write(self.request.user, self.section, obj)
        return can_user_read(self.request.user, self.section, obj)

    def handle_no_permission(self):
        model_name = "record"
        if hasattr(self, "model") and self.model:
            model_name = self.model._meta.verbose_name.lower()

        messages.error(
            self.request,
            f"You do not have permission to view that {model_name}, or it does not exist.",
        )
        return redirect("home")


class DynamicReadPermissionMixin(DynamicPermissionMixin):
    permission_type = "read"


class DynamicWritePermissionMixin(DynamicPermissionMixin):
    permission_type = "write"


class StudentQuerysetRoleMixin:
    """Mixin for views that list Students (or querysets related to a
    Student) which need to be restricted based on the requesting user's
    role: Parents only see their own students, Students only see
    themselves, and other roles see the queryset unrestricted.
    """

    def filter_students_by_role(
        self, qs, adults_field="adults", student_field="pk", empty_queryset=None
    ):
        if empty_queryset is None:
            empty_queryset = Student.objects.none()

        role = get_user_role(self.request.user)
        if role == "Parent":
            try:
                adult = self.request.user.adult_profile
                qs = qs.filter(**{adults_field: adult})
            except (Adult.DoesNotExist, AttributeError):
                qs = empty_queryset
        elif role == "Student":
            try:
                student = self.request.user.student_profile
                value = student.pk if student_field == "pk" else student
                qs = qs.filter(**{student_field: value})
            except (Student.DoesNotExist, AttributeError):
                qs = empty_queryset
        return qs


class LogFormSaveMixin:
    """Mixin to log create/update actions and field changes for ModelForm-based CBVs.

    Logs at INFO level using logger name 'programs.forms'.
    """

    def _fmt_val(self, v):
        try:
            if v is None:
                return "∅"
            s = str(v)
            if len(s) > 200:
                s = s[:200] + "…"
            return s
        except Exception:
            return "<unrepr>"

    def form_valid(self, form):
        model = getattr(form._meta, "model", None)
        model_name = getattr(model, "__name__", form.__class__.__name__)
        user = getattr(getattr(self, "request", None), "user", None)
        user_repr = (
            f"{getattr(user, 'pk', 'anon')}:{getattr(user, 'username', 'anonymous')}"
            if getattr(user, "is_authenticated", False)
            else "anonymous"
        )
        is_create = not bool(getattr(form.instance, "pk", None))

        # Capture changes before saving (for updates)
        changes = []
        try:
            changed_fields = list(getattr(form, "changed_data", []) or [])
            if not is_create and changed_fields and model:
                # Reload from DB to ensure we have current values
                before = model.objects.get(pk=form.instance.pk)
                for f in changed_fields:
                    old = getattr(before, f, None)
                    new = form.cleaned_data.get(f, getattr(form.instance, f, None))
                    if old != new:
                        changes.append((f, old, new))
            elif is_create and changed_fields:
                for f in changed_fields:
                    new = form.cleaned_data.get(f, getattr(form.instance, f, None))
                    changes.append((f, None, new))
        except Exception:
            # Never fail the request due to logging
            changes = []
        except Exception:
            # Never fail the request due to logging
            changes = []

        response = super().form_valid(form)

        obj_id = getattr(
            getattr(self, "object", None), "pk", getattr(form.instance, "pk", None)
        )
        action = "create" if is_create else "update"
        if changes:
            # Map model/action to high-level AuditEvent
            event_map = {
                "Student": AuditEvent.CONTACT_INFO_UPDATED,
                "Adult": AuditEvent.CONTACT_INFO_UPDATED,
                "Enrollment": AuditEvent.ENROLLMENT_CHANGED,
                "AdultStudentRelationship": (
                    AuditEvent.GUARDIAN_ADDED
                    if is_create
                    else AuditEvent.CONTACT_INFO_UPDATED
                ),
            }
            event = event_map.get(str(model_name))
            if event:
                try:
                    before_dict = {
                        str(f): self._fmt_val(old) for f, old, new in changes
                    }
                    after_dict = {str(f): self._fmt_val(new) for f, old, new in changes}
                    log_event(
                        request=getattr(self, "request", None),
                        event=event,
                        resource=form.instance,
                        before=before_dict,
                        after=after_dict,
                        notes=f"{model_name} {action}ed via form save.",
                    )
                except Exception:  # nosec B110
                    # Never crash the main flow for audit logging
                    pass

            for f, old, new in changes:
                forms_logger.info(
                    "FormSave: %s[%s] %s by %s | field=%s | from=%s | to=%s",
                    model_name,
                    obj_id,
                    action,
                    user_repr,
                    f,
                    self._fmt_val(old),
                    self._fmt_val(new),
                )
        else:
            forms_logger.info(
                "FormSave: %s[%s] %s by %s | no field-level differences detected",
                model_name,
                obj_id,
                action,
                user_repr,
            )
        return response


class SortableListViewMixin:
    """Mixin for ListView to support sorting by columns."""

    sort_fields = {}  # Map of field names to actual queryset order_by values
    default_sort_field = None
    default_sort_dir = "asc"

    def get_sort_field(self):
        sort = self.request.GET.get("sort")
        if sort in self.sort_fields:
            return sort
        return self.default_sort_field

    def get_sort_dir(self):
        d = self.request.GET.get("dir", self.default_sort_dir)
        return "desc" if d == "desc" else "asc"

    def apply_sorting(self, queryset):
        sort = self.get_sort_field()
        if not sort:
            return queryset

        direction = self.get_sort_dir()
        order_by_value = self.sort_fields[sort]

        if isinstance(order_by_value, str):
            if direction == "desc":
                if order_by_value.startswith("-"):
                    order_by_value = order_by_value[1:]
                else:
                    order_by_value = f"-{order_by_value}"
            return queryset.order_by(order_by_value)
        elif isinstance(order_by_value, (list, tuple)):
            final_order = []
            for item in order_by_value:
                if isinstance(item, str):
                    if direction == "desc":
                        if item.startswith("-"):
                            final_order.append(item[1:])
                        else:
                            final_order.append(f"-{item}")
                    else:
                        final_order.append(item)
                else:
                    # Handle expressions like Lower('field')
                    if direction == "desc":
                        final_order.append(item.desc())
                    else:
                        final_order.append(item.asc())
            return queryset.order_by(*final_order)
        else:
            # Handle expressions like Lower('field')
            if direction == "desc":
                return queryset.order_by(order_by_value.desc())
            else:
                return queryset.order_by(order_by_value.asc())

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["current_sort"] = self.get_sort_field()
        ctx["current_dir"] = self.get_sort_dir()
        return ctx


class ProgramListView(LoginRequiredMixin, DynamicReadPermissionMixin, ListView):
    model = Program
    template_name = "home.html"  # landing page
    context_object_name = "programs"
    section = "programs"

    def get_queryset(self):
        # Keep a base queryset; ordering will be handled in context via grouping
        qs = Program.objects.all()

        role = get_user_role(self.request.user)
        if role == "Mentor":
            # Only show active programs to Mentors
            from django.db.models import Q
            from django.utils import timezone

            today = timezone.localdate()
            qs = (
                qs.filter(active=True)
                .filter(Q(start_date__isnull=True) | Q(start_date__lte=today))
                .filter(Q(end_date__isnull=True) | Q(end_date__gte=today))
            )
        elif role in ("Student", "Parent", "Alumni"):
            # Students and Parents should not see the program list
            return Program.objects.none()
        return qs

    def get_context_data(self, **kwargs):
        from django.utils import timezone

        ctx = super().get_context_data(**kwargs)

        ctx["role"] = get_user_role(self.request.user)
        today = timezone.localdate()
        programs = list(ctx["programs"])

        def status(prog):
            sd = prog.start_date
            ed = prog.end_date
            if sd and sd > today:
                return "future"
            if ed and ed < today:
                return "past"
            # If only start or only end or none: treat as current if not clearly future/past
            return "current"

        future = sorted(
            [p for p in programs if status(p) == "future"],
            key=lambda p: p.name or "",
        )
        future.sort(
            key=lambda p: (p.start_date is not None, p.start_date), reverse=True
        )

        current = sorted(
            [p for p in programs if status(p) == "current"],
            key=lambda p: p.name or "",
        )
        current.sort(key=lambda p: (p.end_date is not None, p.end_date), reverse=True)

        past = sorted(
            [p for p in programs if status(p) == "past"],
            key=lambda p: p.name or "",
        )
        past.sort(key=lambda p: (p.end_date is not None, p.end_date), reverse=True)

        ctx.update(
            {
                "future_programs": future,
                "current_programs": current,
                "past_programs": past,
            }
        )
        return ctx


class StudentListView(
    LoginRequiredMixin,
    DynamicReadPermissionMixin,
    SortableListViewMixin,
    StudentQuerysetRoleMixin,
    ListView,
):
    model = Student
    template_name = "students/list.html"
    context_object_name = "students"
    section = "student_info"

    sort_fields = {
        "name": (Lower("sort_first"), Lower("last_name")),
        "school": Lower("school__name"),
        "graduation_year": "graduation_year",
        "graduated": "graduated",
    }
    default_sort_field = "name"

    def get_queryset(self):
        qs = super().get_queryset()
        program_id = self.kwargs.get("program_id")
        if program_id:
            qs = qs.filter(enrollment__program_id=program_id).distinct()

        qs = self.filter_students_by_role(qs)

        # Order by preferred/display name if present, otherwise legal first name, then last name (case-insensitive)
        qs = qs.annotate(
            sort_first=Coalesce(NullIf("first_name", Value("")), "legal_first_name"),
        )
        return self.apply_sorting(qs)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        program_id = self.kwargs.get("program_id")
        if program_id:
            ctx["program"] = get_object_or_404(Program, pk=program_id)
        return ctx


class StudentPhotoListView(LoginRequiredMixin, StudentQuerysetRoleMixin, ListView):
    model = Student
    template_name = "students/photo_grid.html"
    context_object_name = "students"
    paginate_by = 48

    def get_queryset(self):
        qs = super().get_queryset()

        qs = self.filter_students_by_role(qs)

        # Order by preferred/display name if present, otherwise legal first name, then last name (case-insensitive)
        return qs.annotate(
            sort_first=Coalesce(NullIf("first_name", Value("")), "legal_first_name"),
        ).order_by(Lower("sort_first"), Lower("last_name"))


class ProgramStudentPhotoListView(
    LoginRequiredMixin, StudentQuerysetRoleMixin, ListView
):
    model = Enrollment
    template_name = "students/photo_grid.html"
    context_object_name = "enrollments"
    paginate_by = 48

    def dispatch(self, request, *args, **kwargs):
        self.program = get_object_or_404(Program, pk=kwargs.get("pk"))
        return super().dispatch(request, *args, **kwargs)

    def get_queryset(self):
        qs = Enrollment.objects.filter(program=self.program).select_related(
            "student", "team"
        )

        qs = self.filter_students_by_role(
            qs,
            adults_field="student__adults",
            student_field="student",
            empty_queryset=Enrollment.objects.none(),
        )

        return qs.annotate(
            sort_first=Lower(
                Coalesce(
                    NullIf("student__first_name", Value("")),
                    "student__legal_first_name",
                )
            ),
            sort_last=Lower("student__last_name"),
        ).order_by("sort_first", "sort_last")

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["program"] = self.program
        # Compatibility for the template which expects 'students'
        ctx["students"] = ctx["enrollments"]
        # Split the page's enrollments into active and inactive sections so
        # inactive (dropped/graduated) students aren't mixed in with active ones.
        page_enrollments = list(ctx["enrollments"])
        ctx["active_enrollments"] = [
            e for e in page_enrollments if e.active and not e.student.graduated
        ]
        ctx["inactive_enrollments"] = [
            e for e in page_enrollments if not (e.active and not e.student.graduated)
        ]
        return ctx


class StudentEmergencyContactsView(
    LoginRequiredMixin, SortableListViewMixin, StudentQuerysetRoleMixin, ListView
):
    model = Student
    template_name = "students/emergency_contacts.html"
    context_object_name = "students"

    sort_fields = {
        "name": (Lower("sort_first"), Lower("last_name")),
        "school": Lower("school__name"),
    }
    default_sort_field = "name"

    def get_queryset(self):
        qs = super().get_queryset().filter(graduated=False)

        qs = self.filter_students_by_role(qs)

        qs = (
            qs.select_related("school", "primary_contact", "secondary_contact")
            .prefetch_related("adults")
            .annotate(
                sort_first=Coalesce(
                    NullIf("first_name", Value("")), "legal_first_name"
                ),
            )
        )
        return self.apply_sorting(qs)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        # Backwards compatibility: some templates expect 'plist'
        ctx.setdefault("plist", ctx.get("students") or ctx.get("object_list"))
        return ctx


class StudentsByGradeView(LoginRequiredMixin, StudentQuerysetRoleMixin, ListView):
    model = Student
    template_name = "students/by_grade.html"
    context_object_name = "students"

    def get_queryset(self):
        qs = super().get_queryset().filter(graduated=False)

        qs = self.filter_students_by_role(qs)

        return (
            qs.select_related("school")
            .annotate(
                sort_first=Coalesce(
                    NullIf("first_name", Value("")), "legal_first_name"
                ),
            )
            .order_by("graduation_year", Lower("sort_first"), Lower("last_name"))
        )

    def get_context_data(self, **kwargs):
        from django.utils import timezone

        ctx = super().get_context_data(**kwargs)
        current_year = timezone.now().year

        def compute_grade(gy):
            if not gy:
                return None
            # Approximate US grade level: graduating this year = 12th grade now
            return 12 - (gy - current_year)

        def label_for_grade(grade_num):
            if grade_num is None:
                return "Unknown Grade"
            if grade_num < 0:
                return "Pre-K"
            if grade_num == 0:
                return "Kindergarten"
            # Ordinal suffixes
            n = int(grade_num)
            if 10 <= (n % 100) <= 13:
                suffix = "th"
            else:
                suffix = {1: "st", 2: "nd", 3: "rd"}.get(n % 10, "th")
            return f"{n}{suffix} Grade"

        grouped = {}
        grade_order = {}
        for s in ctx["students"]:
            g = compute_grade(getattr(s, "graduation_year", None))
            label = label_for_grade(g)
            grouped.setdefault(label, []).append(s)
            if label not in grade_order:
                grade_order[label] = -1000 if g is None else int(g)
        # Sort labels by grade number descending (12 -> 11 -> ... -> 1 -> 0 -> negatives), Unknown last
        sorted_labels = sorted(
            grade_order.keys(), key=lambda lbl: grade_order[lbl], reverse=True
        )
        # Ensure 'Unknown Grade' is at the very end
        if "Unknown Grade" in sorted_labels:
            sorted_labels = [lbl for lbl in sorted_labels if lbl != "Unknown Grade"] + [
                "Unknown Grade"
            ]
        ctx["grouped"] = [(label, grouped.get(label, [])) for label in sorted_labels]
        return ctx


class StudentsBySchoolView(LoginRequiredMixin, ListView):
    model = Student
    template_name = "students/by_school.html"
    context_object_name = "students"

    def get_queryset(self):
        qs = super().get_queryset().filter(graduated=False)

        role = get_user_role(self.request.user)
        if role == "Parent":
            try:
                adult = self.request.user.adult_profile
                qs = qs.filter(adults=adult)
            except (Adult.DoesNotExist, AttributeError):
                qs = Student.objects.none()
        elif role == "Student":
            try:
                student = self.request.user.student_profile
                qs = qs.filter(pk=student.pk)
            except (Student.DoesNotExist, AttributeError):
                qs = Student.objects.none()

        return (
            qs.select_related("school")
            .annotate(
                sort_first=Coalesce(
                    NullIf("first_name", Value("")), "legal_first_name"
                ),
            )
            .order_by("school__name", Lower("sort_first"), Lower("last_name"))
        )

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)

        ctx["role"] = get_user_role(self.request.user)
        grouped = {}
        for s in ctx["students"]:
            label = s.school.name if s.school_id else "No School"
            grouped.setdefault(label, []).append(s)
        # Sort by school label
        ctx["grouped"] = sorted(
            grouped.items(), key=lambda kv: (kv[0] == "No School", kv[0])
        )
        return ctx


class ParentListView(LoginRequiredMixin, SortableListViewMixin, ListView):
    model = Adult
    template_name = "parents/list.html"
    context_object_name = "parents"

    sort_fields = {
        "name": (Lower("first_name"), Lower("last_name")),
        "email": Lower("personal_email"),
        "phone": "phone_number",
    }
    default_sort_field = "name"

    def get_queryset(self):
        qs = Adult.objects.filter(is_parent=True).prefetch_related("students")
        program_id = self.kwargs.get("program_id")
        if program_id:
            qs = qs.filter(students__enrollment__program_id=program_id).distinct()

        role = get_user_role(self.request.user)
        if role == "Parent":
            try:
                adult = self.request.user.adult_profile
                student_ids = adult.students.values_list("id", flat=True)
                qs = qs.filter(students__id__in=student_ids).distinct()
            except:
                qs = Adult.objects.none()
        elif role == "Student":
            try:
                student = self.request.user.student_profile
                qs = qs.filter(students=student).distinct()
            except:
                qs = Adult.objects.none()

        return self.apply_sorting(qs)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        program_id = self.kwargs.get("program_id")
        if program_id:
            ctx["program"] = get_object_or_404(Program, pk=program_id)
        return ctx


class MentorListView(LoginRequiredMixin, SortableListViewMixin, ListView):
    model = Adult
    template_name = "mentors/list.html"
    context_object_name = "mentors"

    sort_fields = {
        "name": (Lower("first_name"), Lower("last_name")),
        "role": "role",
        "active": "active",
    }
    default_sort_field = "name"

    def get_queryset(self):
        qs = Adult.objects.filter(is_mentor=True)
        program_id = self.kwargs.get("program_id")
        if program_id:
            qs = qs.filter(students__enrollment__program_id=program_id).distinct()
        return self.apply_sorting(qs)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        program_id = self.kwargs.get("program_id")
        if program_id:
            ctx["program"] = get_object_or_404(Program, pk=program_id)
        return ctx


class AlumniListView(LoginRequiredMixin, SortableListViewMixin, ListView):
    model = Adult
    template_name = "alumni/list.html"
    context_object_name = "alumni"

    sort_fields = {
        "name": (Lower("first_name"), Lower("last_name")),
        "email": Lower("personal_email"),
        "phone": "phone_number",
        "college": Lower("college"),
        "employer": Lower("employer"),
        "ok_to_contact": "ok_to_contact",
    }
    default_sort_field = "name"

    def get_queryset(self):
        qs = Adult.objects.filter(is_alumni=True)
        program_id = self.kwargs.get("program_id")
        if program_id:
            qs = qs.filter(students__enrollment__program_id=program_id).distinct()
        return self.apply_sorting(qs)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        program_id = self.kwargs.get("program_id")
        if program_id:
            ctx["program"] = get_object_or_404(Program, pk=program_id)
        return ctx


class StudentConvertToAlumniView(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = "programs.change_student"

    def post(self, request, pk):
        from .utils import convert_student_to_alumni

        student = get_object_or_404(Student, pk=pk)
        _adult, created, _marked = convert_student_to_alumni(student)
        if created:
            messages.success(
                request,
                f"Converted {student} to Alumni (Adult created) and marked student as graduated.",
            )
        else:
            messages.info(
                request,
                f"{student} is now marked as Alumni. Student marked as graduated.",
            )
        return redirect_back(request, "student_list")


class StudentBulkConvertToAlumniView(LoginRequiredMixin, LeadMentorRequiredMixin, View):
    template_name = "students/convert_to_alumni.html"

    def get(self, request):
        from django.utils import timezone

        year = request.GET.get("year")
        try:
            year = int(year) if year else timezone.now().year
        except ValueError:
            year = timezone.now().year
        # Default to seniors: graduation_year equals the selected year, and active (non-graduated)
        students = (
            Student.objects.filter(graduation_year=year, graduated=False)
            .annotate(
                sort_first=Coalesce(NullIf("first_name", Value("")), "legal_first_name")
            )
            .order_by(Lower("sort_first"), Lower("last_name"))
        )
        return render(
            request,
            self.template_name,
            {
                "year": year,
                "students": students,
            },
        )

    def post(self, request):
        from .utils import convert_student_to_alumni, find_matching_alumni_adult

        action = request.POST.get("action", "convert")
        ids = request.POST.getlist("student_ids")
        year = request.POST.get("year")

        # Validate and normalize year
        try:
            if year:
                year = str(int(year))
        except (ValueError, TypeError):
            year = None

        if not ids:
            messages.info(request, "No students selected.")
            if year:
                base_url = reverse("student_bulk_convert_select")
                query = QueryDict("", mutable=True)
                query["year"] = year
                return HttpResponseRedirect(f"{base_url}?{query.urlencode()}")
            return redirect("student_bulk_convert_select")

        qs = Student.objects.filter(pk__in=ids).order_by("last_name", "first_name")

        if action == "preview":
            # Build preview info without writing changes against Adults flagged as alumni
            will_create = []
            already_alumni = []
            for s in qs:
                if find_matching_alumni_adult(s):
                    already_alumni.append(s)
                else:
                    will_create.append(s)
            will_mark_graduated = [s for s in qs if not s.graduated]
            return render(
                request,
                "students/convert_to_alumni_preview.html",
                {
                    "year": year,
                    "students": qs,
                    "will_create": will_create,
                    "already_alumni": already_alumni,
                    "will_mark_graduated": will_mark_graduated,
                    "ids": ids,
                },
            )

        # Default: perform conversion
        created = 0
        existed = 0
        marked_graduated = 0

        for student in qs:
            _adult, was_created, was_marked = convert_student_to_alumni(student)
            if was_created:
                created += 1
            else:
                existed += 1
            if was_marked:
                marked_graduated += 1
        messages.success(
            request,
            f"Converted {created} new alumni (Adults), {existed} already existed/updated. "
            f"Marked {marked_graduated} student(s) as graduated.",
        )
        return redirect("alumni_list")


class ImportDashboardView(LoginRequiredMixin, View):
    def get(self, request):
        from django.shortcuts import render

        from .models import Program

        programs = Program.objects.all().order_by("name")
        programs_with_attendance = [p for p in programs if p.has_feature("attendance")]
        return render(
            request,
            "imports/dashboard.html",
            {
                "programs": programs,
                "attendance_programs": programs_with_attendance,
            },
        )


class StudentImportView(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = "programs.add_student"

    def post(self, request):
        file = request.FILES.get("file")
        if not file:
            messages.error(request, "No file uploaded.")
            return redirect_back(request, "import_dashboard")
        name = file.name.lower()
        overwrite = request.POST.get("overwrite") == "1"
        created = 0
        updated = 0
        errors = 0
        try:
            if name.endswith(".csv"):
                import csv
                import io

                text = io.TextIOWrapper(file.file, encoding="utf-8")
                reader = csv.DictReader(text)
                rows = list(reader)
            elif name.endswith(".xlsx"):
                from openpyxl import load_workbook

                wb = load_workbook(filename=file, read_only=True, data_only=True)
                ws = wb.active
                headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
                rows = []
                for r in ws.iter_rows(min_row=2, values_only=True):
                    rows.append(
                        {
                            str(headers[i]): r[i]
                            for i in range(len(headers))
                            if headers[i] is not None
                        }
                    )
            else:
                messages.error(
                    request, "Unsupported file type. Please upload CSV or XLSX."
                )
                return redirect("import_dashboard")

            # Helpers
            from datetime import date, datetime

            def raw(d, *keys):
                for k in keys:
                    if k in d and d[k] is not None:
                        return d[k]
                return None

            def val(d, *keys):
                for k in keys:
                    if k in d and d[k] is not None:
                        v = str(d[k]).strip()
                        if v != "" and v.lower() != "none":
                            return v
                return None

            def val_bool(d, *keys):
                v = val(d, *keys)
                if v is None:
                    return None
                s = v.strip().lower()
                if s in ("y", "yes", "true", "t", "1"):
                    return True
                if s in ("n", "no", "false", "f", "0"):
                    return False
                return None

            def val_date(d, *keys):
                # Accept date objects from XLSX or parse common string formats
                rv = raw(d, *keys)
                if isinstance(rv, datetime):
                    return rv.date()
                if isinstance(rv, date):
                    return rv
                v = val(d, *keys)
                if not v:
                    return None
                for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
                    try:
                        return datetime.strptime(v, fmt).date()
                    except ValueError:
                        continue
                return None

            def get_or_create_parent(first, last, email):
                # Try to find by email first
                if email:
                    p = Adult.objects.filter(personal_email__iexact=email).first()
                    if p:
                        if overwrite:
                            changed_parent = False
                            if first and p.first_name != first:
                                p.first_name = first
                                changed_parent = True
                            if last and p.last_name != last:
                                p.last_name = last
                                changed_parent = True
                            if changed_parent:
                                p.save()
                        return p
                # Next try by name match
                if first and last:
                    p = Adult.objects.filter(
                        first_name__iexact=first, last_name__iexact=last
                    ).first()
                    if p:
                        if (
                            overwrite
                            and email
                            and (p.personal_email or "").lower()
                            != (email or "").lower()
                        ):
                            p.personal_email = email
                            p.save()
                        return p
                # If we have at least one of name or email, create
                if first or last or email:
                    return Adult.objects.create(
                        first_name=first
                        or (email.split("@")[0] if email else "Parent"),
                        last_name=last or "(contact)",
                        personal_email=email or None,
                        is_parent=True,
                    )
                return None

            for d in rows:
                first = val(d, "first_name", "First Name", "Preferred First Name")
                legal_first = val(d, "legal_first_name", "Legal First Name") or first
                last = val(d, "last_name", "Last Name")
                if not last or not legal_first:
                    errors += 1
                    continue

                # Simple strings
                pronouns = val(d, "pronouns", "Pronouns")
                address = val(d, "address", "Address", "Street Address")
                city = val(d, "city", "City")
                state = val(d, "state", "State")
                zip_code = val(d, "zip_code", "Zip Code", "ZIP", "Zip")
                cell_phone = val(
                    d,
                    "cell_phone_number",
                    "Cell Phone Number",
                    "Cell Phone",
                    "Phone",
                    "Phone Number",
                )
                personal_email = val(d, "personal_email", "Email", "Personal Email")
                andrew_id = val(d, "andrew_id", "Andrew ID", "AndrewID")
                andrew_email = val(d, "andrew_email", "Andrew Email")
                race_ethnicity = val(
                    d, "race_ethnicity", "Race/Ethnicity", "Race", "Ethnicity"
                )
                tshirt_size = val(d, "tshirt_size", "T-Shirt Size", "Shirt Size")
                discord_handle = val(
                    d, "discord_handle", "Discord Handle", "Discord", "Discord Username"
                )

                # Dates and booleans
                dob = val_date(d, "date_of_birth", "Date of Birth", "DOB", "Birthdate")
                seen_once = val_bool(d, "seen_once", "Seen Once")
                on_discord = val_bool(d, "on_discord", "On Discord")
                graduated = val_bool(d, "graduated", "Graduated")
                # Backward compatibility for older templates that still send Active.
                # Student now uses "graduated" instead of an "active" field.
                active = val_bool(d, "active", "Active")
                if graduated is None and active is not None:
                    graduated = not active

                # School/year
                school_name = val(d, "school", "School")
                grad = val(d, "graduation_year", "Graduation Year")
                school = None
                if school_name:
                    school, _ = School.objects.get_or_create(name=school_name)
                grad_year = None
                if grad and str(grad).isdigit():
                    grad_year = int(str(grad))

                obj, created_flag = Student.objects.get_or_create(
                    last_name=last,
                    legal_first_name=legal_first,
                    defaults={
                        "first_name": first if first != legal_first else None,
                        "pronouns": pronouns,
                        "date_of_birth": dob,
                        "address": address,
                        "city": city,
                        "state": state,
                        "zip_code": zip_code,
                        "phone_number": cell_phone,
                        "phone_type": "cell",
                        "can_receive_texts": True,
                        "personal_email": personal_email,
                        "andrew_id": andrew_id,
                        "andrew_email": andrew_email,
                        "tshirt_size": tshirt_size,
                        "seen_once": seen_once if seen_once is not None else False,
                        "on_discord": on_discord if on_discord is not None else False,
                        "discord_handle": discord_handle,
                        "school": school,
                        "graduation_year": grad_year,
                        "graduated": graduated if graduated is not None else False,
                    },
                )
                if created_flag:
                    created += 1
                elif overwrite:
                    changed = False
                    # Strings and relations
                    for field, value in [
                        ("first_name", first),
                        ("pronouns", pronouns),
                        ("address", address),
                        ("city", city),
                        ("state", state),
                        ("zip_code", zip_code),
                        ("phone_number", cell_phone),
                        ("phone_type", "cell"),
                        ("can_receive_texts", True),
                        ("personal_email", personal_email),
                        ("andrew_id", andrew_id),
                        ("andrew_email", andrew_email),
                        ("tshirt_size", tshirt_size),
                        ("discord_handle", discord_handle),
                    ]:
                        if value and getattr(obj, field) != value:
                            setattr(obj, field, value)
                            changed = True
                    if dob and obj.date_of_birth != dob:
                        obj.date_of_birth = dob
                        changed = True
                    if school and obj.school != school:
                        obj.school = school
                        changed = True
                    if grad_year and obj.graduation_year != grad_year:
                        obj.graduation_year = grad_year
                        changed = True
                    # Booleans (allow False updates)
                    if seen_once is not None and obj.seen_once != seen_once:
                        obj.seen_once = seen_once
                        changed = True
                    if on_discord is not None and obj.on_discord != on_discord:
                        obj.on_discord = on_discord
                        changed = True
                    if graduated is not None and obj.graduated != graduated:
                        obj.graduated = graduated
                        changed = True
                    if changed:
                        obj.save()
                        updated += 1

                # Map race/ethnicity text to multi-select options
                try:
                    opts = RaceEthnicity.match_from_text(race_ethnicity)
                    if opts.exists():
                        obj.race_ethnicities.set(list(opts))
                except Exception:
                    logger.debug(
                        "Race/Ethnicity matching failed during import", exc_info=True
                    )

                # Parent linkage (primary and secondary)
                prim_first = val(
                    d,
                    "primary_parent_first_name",
                    "Primary Parent First Name",
                    "Primary First Name",
                    "Primary First",
                )
                prim_last = val(
                    d,
                    "primary_parent_last_name",
                    "Primary Parent Last Name",
                    "Primary Last Name",
                    "Primary Last",
                )
                prim_email = val(
                    d,
                    "primary_parent_email",
                    "Primary Parent Email",
                    "Primary Email",
                    "Primary E-mail",
                    "Primary Email Address",
                )
                sec_first = val(
                    d,
                    "secondary_parent_first_name",
                    "Secondary Parent First Name",
                    "Secondary First Name",
                    "Secondary First",
                )
                sec_last = val(
                    d,
                    "secondary_parent_last_name",
                    "Secondary Parent Last Name",
                    "Secondary Last Name",
                    "Secondary Last",
                )
                sec_email = val(
                    d,
                    "secondary_parent_email",
                    "Secondary Parent Email",
                    "Secondary Email",
                    "Secondary E-mail",
                    "Secondary Email Address",
                )

                contact_changed = False
                primary = get_or_create_parent(prim_first, prim_last, prim_email)
                secondary = get_or_create_parent(sec_first, sec_last, sec_email)
                if primary:
                    if obj.primary_contact_id != getattr(primary, "id", None):
                        obj.primary_contact = primary
                        contact_changed = True
                    # Ensure M2M link exists (both sides)
                    if primary.id and not obj.adults.filter(id=primary.id).exists():
                        obj.adults.add(primary)
                        primary.students.add(obj)
                if secondary:
                    if obj.secondary_contact_id != getattr(secondary, "id", None):
                        obj.secondary_contact = secondary
                        contact_changed = True
                    if secondary.id and not obj.adults.filter(id=secondary.id).exists():
                        obj.adults.add(secondary)
                        secondary.students.add(obj)
                if contact_changed:
                    obj.save(
                        update_fields=[
                            "primary_contact",
                            "secondary_contact",
                            "updated_at",
                        ]
                    )
                    if not created_flag:
                        # Only count as updated when not newly created and not already counted
                        updated += 1
            if created or updated:
                messages.success(
                    request,
                    f"Imported {created} new, updated {updated}. Skipped {errors}.",
                )
            else:
                messages.info(request, "No rows imported.")
        except Exception as e:
            messages.error(request, f"Import failed: {e}")
        return redirect_back(request, "import_dashboard")


class ParentImportView(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = "programs.add_adult"

    def post(self, request):
        file = request.FILES.get("file")
        if not file:
            messages.error(request, "No file uploaded.")
            return redirect_back(request, "import_dashboard")
        name = file.name.lower()
        overwrite = request.POST.get("overwrite") == "1"
        created = 0
        updated = 0
        errors = 0
        try:
            if name.endswith(".csv"):
                import csv
                import io

                text = io.TextIOWrapper(file.file, encoding="utf-8")
                reader = csv.DictReader(text)
                rows = list(reader)
            elif name.endswith(".xlsx"):
                from openpyxl import load_workbook

                wb = load_workbook(filename=file, read_only=True)
                ws = wb.active
                headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
                rows = []
                for r in ws.iter_rows(min_row=2, values_only=True):
                    rows.append(
                        {
                            str(headers[i]): r[i]
                            for i in range(len(headers))
                            if headers[i] is not None
                        }
                    )
            else:
                messages.error(
                    request, "Unsupported file type. Please upload CSV or XLSX."
                )
                return redirect("import_dashboard")

            def val(d, *keys):
                for k in keys:
                    if k in d and d[k] is not None:
                        v = str(d[k]).strip()
                        if v != "" and v.lower() != "none":
                            return v
                return None

            for d in rows:
                first = val(d, "first_name", "First Name")
                last = val(d, "last_name", "Last Name")
                if not first or not last:
                    errors += 1
                    continue
                email = val(d, "email", "Email")
                phone = val(
                    d,
                    "cell_phone",
                    "Cell Phone",
                    "Cell Phone Number",
                    "Phone",
                    "Phone Number",
                )
                obj, created_flag = Adult.objects.get_or_create(
                    first_name=first,
                    last_name=last,
                    defaults={
                        "personal_email": email,
                        "phone_number": phone,
                        "phone_type": "cell",
                        "can_receive_texts": True,
                        "is_parent": True,
                    },
                )
                if created_flag:
                    created += 1
                else:
                    changed = False
                    if not obj.is_parent:
                        obj.is_parent = True
                        changed = True

                    if not overwrite:
                        if changed:
                            obj.save(update_fields=["is_parent", "updated_at"])
                            updated += 1
                        continue

                    if email and obj.personal_email != email:
                        obj.personal_email = email
                        changed = True
                    if phone and obj.phone_number != phone:
                        obj.phone_number = phone
                        obj.phone_type = "cell"
                        obj.can_receive_texts = True
                        changed = True
                    if changed:
                        obj.save()
                        updated += 1
            messages.success(
                request, f"Imported {created} new, updated {updated}. Skipped {errors}."
            )
        except Exception as e:
            messages.error(request, f"Import failed: {e}")
        return redirect_back(request, "import_dashboard")


class RelationshipImportView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """
    Re-link existing Students to Parent/Adult records and set relationship types.
    Safe to run multiple times (idempotent). Optionally supports dry-run.
    """

    permission_required = "programs.change_student"

    def post(self, request):
        file = request.FILES.get("file")
        if not file:
            messages.error(request, "No file uploaded.")
            return redirect_back(request, "import_dashboard")
        name = file.name.lower()
        dry_run = request.POST.get("dry_run") in ("1", "on", "true", "True")
        overwrite = request.POST.get("overwrite") == "1"
        can_create_parents = request.user.has_perm("programs.add_adult")

        linked = 0
        set_primary = 0
        set_secondary = 0
        rel_updated = 0
        created_parents = 0
        would_create_parents = 0
        missing_or_ambiguous_students = 0
        skipped = 0

        try:
            # Parse CSV/XLSX similar to other imports
            if name.endswith(".csv"):
                import csv
                import io

                text = io.TextIOWrapper(file.file, encoding="utf-8")
                reader = csv.DictReader(text)
                rows = list(reader)
            elif name.endswith(".xlsx"):
                from openpyxl import load_workbook

                wb = load_workbook(filename=file, read_only=True, data_only=True)
                ws = wb.active
                headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
                rows = []
                for r in ws.iter_rows(min_row=2, values_only=True):
                    rows.append(
                        {
                            str(headers[i]): r[i]
                            for i in range(len(headers))
                            if headers[i] is not None
                        }
                    )
            else:
                messages.error(
                    request, "Unsupported file type. Please upload CSV or XLSX."
                )
                return redirect("import_dashboard")

            # Helpers
            from datetime import date, datetime

            def raw(d, *keys):
                for k in keys:
                    if k in d and d[k] is not None:
                        return d[k]
                return None

            def val(d, *keys):
                for k in keys:
                    if k in d and d[k] is not None:
                        v = str(d[k]).strip()
                        if v != "" and v.lower() != "none":
                            return v
                return None

            def val_date(d, *keys):
                rv = raw(d, *keys)
                if isinstance(rv, datetime):
                    return rv.date()
                if isinstance(rv, date):
                    return rv
                s = val(d, *keys)
                if not s:
                    return None
                for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
                    try:
                        return datetime.strptime(s, fmt).date()
                    except ValueError:
                        continue
                return None

            def normalize_rel(s):
                if not s:
                    return None
                s2 = s.strip().lower()
                # Accept either key or display label
                keys = {k for k, _ in RELATIONSHIP_CHOICES}
                labels = {lbl.lower(): k for k, lbl in RELATIONSHIP_CHOICES}
                synonyms = {
                    "mom": "mother",
                    "dad": "father",
                    "grandma": "grandmother",
                    "grandpa": "grandfather",
                    "guardian": "guardian",
                    "parent": "parent",
                }
                if s2 in keys:
                    return s2
                if s2 in labels:
                    return labels[s2]
                if s2 in synonyms and synonyms[s2] in keys:
                    return synonyms[s2]
                return None

            def resolve_student(d):
                # Priority: ID -> Andrew ID -> (First/Legal First + Last + DOB) -> (First/Legal First + Last)
                sid = val(d, "student_id", "Student ID", "ID")
                if sid and str(sid).isdigit():
                    st = Student.objects.filter(pk=int(str(sid))).first()
                    if st:
                        return st

                aid = val(d, "andrew_id", "Andrew ID", "AndrewID")
                if aid:
                    st = Student.objects.filter(andrew_id__iexact=aid).first()
                    if st:
                        return st

                last = val(d, "last_name", "Last Name")
                first = val(d, "first_name", "First Name", "Preferred First Name")
                legal_first = val(d, "legal_first_name", "Legal First Name") or first
                dob = val_date(d, "date_of_birth", "Date of Birth", "DOB", "Birthdate")

                if not last or not legal_first:
                    return None

                qs = Student.objects.filter(
                    last_name__iexact=last, legal_first_name__iexact=legal_first
                )
                if dob:
                    qs = qs.filter(date_of_birth=dob)
                count = qs.count()
                if count == 1:
                    return qs.first()
                if count == 0 and first and first != legal_first:
                    # Try match on preferred first + last (+dob)
                    qs = Student.objects.filter(
                        last_name__iexact=last, first_name__iexact=first
                    )
                    if dob:
                        qs = qs.filter(date_of_birth=dob)
                    if qs.count() == 1:
                        return qs.first()
                return None if qs.count() != 1 else qs.first()

            def find_or_create_parent(first, last, email):
                # Try resolve by email first
                p = None
                if email:
                    p = Adult.objects.filter(personal_email__iexact=email).first()
                if not p and first and last:
                    p = Adult.objects.filter(
                        first_name__iexact=first, last_name__iexact=last
                    ).first()
                created = False
                if not p:
                    if dry_run or not can_create_parents:
                        return None, False, True  # would create
                    p = Adult.objects.create(
                        first_name=first
                        or (email.split("@")[0] if email else "Parent"),
                        last_name=last or "(contact)",
                        personal_email=email or None,
                        is_parent=True,
                    )
                    created = True
                else:
                    # If we found existing Adult but not flagged as parent, set it
                    if not dry_run and not p.is_parent:
                        p.is_parent = True
                        p.save(update_fields=["is_parent", "updated_at"])
                return p, created, False

            for d in rows:
                student = resolve_student(d)
                if not student:
                    missing_or_ambiguous_students += 1
                    continue

                groups = [
                    {
                        "role": "primary",
                        "first": val(
                            d,
                            "primary_parent_first_name",
                            "Primary Parent First Name",
                            "Primary First Name",
                            "Primary First",
                        ),
                        "last": val(
                            d,
                            "primary_parent_last_name",
                            "Primary Parent Last Name",
                            "Primary Last Name",
                            "Primary Last",
                        ),
                        "email": val(
                            d,
                            "primary_parent_email",
                            "Primary Parent Email",
                            "Primary Email",
                        ),
                        "rel": val(
                            d,
                            "primary_parent_relationship",
                            "Primary Parent Relationship",
                            "Primary Relationship",
                        ),
                    },
                    {
                        "role": "secondary",
                        "first": val(
                            d,
                            "secondary_parent_first_name",
                            "Secondary Parent First Name",
                            "Secondary First Name",
                            "Secondary First",
                        ),
                        "last": val(
                            d,
                            "secondary_parent_last_name",
                            "Secondary Parent Last Name",
                            "Secondary Last Name",
                            "Secondary Last",
                        ),
                        "email": val(
                            d,
                            "secondary_parent_email",
                            "Secondary Parent Email",
                            "Secondary Email",
                        ),
                        "rel": val(
                            d,
                            "secondary_parent_relationship",
                            "Secondary Parent Relationship",
                            "Secondary Relationship",
                        ),
                    },
                ]

                updated_student_fields = set()

                for g in groups:
                    if not (g["first"] or g["last"] or g["email"]):
                        continue
                    adult, created_flag, would_create = find_or_create_parent(
                        g["first"], g["last"], g["email"]
                    )
                    if would_create:
                        would_create_parents += 1
                        continue
                    if created_flag:
                        created_parents += 1
                    if not adult:
                        skipped += 1
                        continue

                    # Relationship type
                    rel_key = normalize_rel(g["rel"])
                    rel_key = rel_key or "parent"
                    if not dry_run:
                        _, rel_created = (
                            AdultStudentRelationship.objects.update_or_create(
                                adult=adult,
                                student=student,
                                defaults={"relationship_to_student": rel_key},
                            )
                        )
                        if rel_created:
                            linked += 1
                    else:
                        if not student.adults.filter(id=adult.id).exists():
                            linked += 1
                    rel_updated += 1

                    # Ensure Adult is linked to Student (M2M) - handled by update_or_create when not dry_run.

                    # Optionally set primary/secondary contact
                    if g["role"] == "primary":
                        if student.primary_contact_id != adult.id:
                            if not dry_run and overwrite:
                                student.primary_contact = adult
                                updated_student_fields.add("primary_contact")
                            set_primary += 1
                    elif g["role"] == "secondary":
                        if student.secondary_contact_id != adult.id:
                            if not dry_run and overwrite:
                                student.secondary_contact = adult
                                updated_student_fields.add("secondary_contact")
                            set_secondary += 1

                if updated_student_fields and not dry_run:
                    fields = list(updated_student_fields) + ["updated_at"]
                    student.save(update_fields=fields)

            # Compose message
            notes = []
            if dry_run:
                notes.append("DRY RUN (no changes saved)")
            if not can_create_parents:
                notes.append(
                    "Note: lacking permission to create parents; rows requiring new parent were skipped."
                )
            extras = f" {'; '.join(notes)}" if notes else ""
            messages.success(
                request,
                f"Relationships import: linked {linked} (primary set {set_primary}, secondary set {set_secondary}); "
                f"updated relationship types {rel_updated}; "
                f"created parents {created_parents}"
                f"{(' (would create: ' + str(would_create_parents) + ')' if dry_run else '')}; "
                f"missing/ambiguous students {missing_or_ambiguous_students}; skipped {skipped}.{extras}",
            )
        except Exception as e:
            messages.error(request, f"Import failed: {e}")
        return redirect_back(request, "import_dashboard")


class MentorImportView(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = "programs.add_adult"

    def post(self, request):
        file = request.FILES.get("file")
        if not file:
            messages.error(request, "No file uploaded.")
            return redirect_back(request, "import_dashboard")
        name = file.name.lower()
        overwrite = request.POST.get("overwrite") == "1"
        created = 0
        updated = 0
        errors = 0
        try:
            if name.endswith(".csv"):
                import csv
                import io

                text = io.TextIOWrapper(file.file, encoding="utf-8")
                reader = csv.DictReader(text)
                rows = list(reader)
            elif name.endswith(".xlsx"):
                from openpyxl import load_workbook

                wb = load_workbook(filename=file, read_only=True)
                ws = wb.active
                headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
                rows = []
                for r in ws.iter_rows(min_row=2, values_only=True):
                    rows.append(
                        {
                            str(headers[i]): r[i]
                            for i in range(len(headers))
                            if headers[i] is not None
                        }
                    )
            else:
                messages.error(
                    request, "Unsupported file type. Please upload CSV or XLSX."
                )
                return redirect("import_dashboard")

            def val(d, *keys):
                for k in keys:
                    if k in d and d[k] is not None:
                        v = str(d[k]).strip()
                        if v != "" and v.lower() != "none":
                            return v
                return None

            for d in rows:
                first = val(d, "first_name", "First Name")
                last = val(d, "last_name", "Last Name")
                if not first or not last:
                    errors += 1
                    continue
                email = val(d, "personal_email", "Email", "Personal Email")
                andrew_email = val(d, "andrew_email", "Andrew Email")
                role = val(d, "role", "Role") or "mentor"
                obj, created_flag = Adult.objects.get_or_create(
                    first_name=first,
                    last_name=last,
                    defaults={
                        "personal_email": email,
                        "andrew_email": andrew_email,
                        "role": role,
                        "is_mentor": True,
                    },
                )
                if created_flag:
                    created += 1
                else:
                    changed = False
                    if not obj.is_mentor:
                        obj.is_mentor = True
                        changed = True

                    if not overwrite:
                        if changed:
                            obj.save(update_fields=["is_mentor", "updated_at"])
                            updated += 1
                        continue

                    for field, value in [
                        ("personal_email", email),
                        ("andrew_email", andrew_email),
                        ("role", role),
                    ]:
                        if value and getattr(obj, field) != value:
                            setattr(obj, field, value)
                            changed = True
                    if changed:
                        obj.save()
                        updated += 1
            messages.success(
                request, f"Imported {created} new, updated {updated}. Skipped {errors}."
            )
        except Exception as e:
            messages.error(request, f"Import failed: {e}")
        return redirect_back(request, "import_dashboard")


class SchoolImportView(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = "programs.add_school"

    def post(self, request):
        file = request.FILES.get("file")
        if not file:
            messages.error(request, "No file uploaded.")
            return redirect_back(request, "import_dashboard")
        name = file.name.lower()
        overwrite = request.POST.get("overwrite") == "1"
        created = 0
        updated = 0
        errors = 0
        try:
            if name.endswith(".csv"):
                import csv
                import io

                text = io.TextIOWrapper(file.file, encoding="utf-8")
                reader = csv.DictReader(text)
                rows = list(reader)
            elif name.endswith(".xlsx"):
                from openpyxl import load_workbook

                wb = load_workbook(filename=file, read_only=True)
                ws = wb.active
                headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
                rows = []
                for r in ws.iter_rows(min_row=2, values_only=True):
                    rows.append(
                        {
                            str(headers[i]): r[i]
                            for i in range(len(headers))
                            if headers[i] is not None
                        }
                    )
            else:
                messages.error(
                    request, "Unsupported file type. Please upload CSV or XLSX."
                )
                return redirect("import_dashboard")

            def val(d, *keys):
                for k in keys:
                    if k in d and d[k] is not None:
                        v = str(d[k]).strip()
                        if v != "" and v.lower() != "none":
                            return v
                return None

            for d in rows:
                school_name = val(d, "name", "Name", "School")
                if not school_name:
                    errors += 1
                    continue
                district = val(d, "district", "District", "School District")
                street = val(d, "street_address", "Street", "Street Address", "Address")
                city = val(d, "city", "City")
                state = val(d, "state", "State")
                zip_code = val(
                    d, "zip", "ZIP", "Zip", "zip_code", "Zip Code", "Postal Code"
                )
                obj, created_flag = School.objects.get_or_create(
                    name=school_name,
                    defaults={
                        "district": district,
                        "street_address": street,
                        "city": city,
                        "state": state,
                        "zip_code": zip_code,
                    },
                )
                if created_flag:
                    created += 1
                elif overwrite:
                    changed = False
                    for field, value in [
                        ("district", district),
                        ("street_address", street),
                        ("city", city),
                        ("state", state),
                        ("zip_code", zip_code),
                    ]:
                        if value and getattr(obj, field) != value:
                            setattr(obj, field, value)
                            changed = True
                    if changed:
                        obj.save()
                        updated += 1
            messages.success(
                request, f"Imported {created} new, updated {updated}. Skipped {errors}."
            )
        except Exception as e:
            messages.error(request, f"Import failed: {e}")
        return redirect_back(request, "import_dashboard")


class MentorCreateView(
    PassUserToFormMixin,
    LogFormSaveMixin,
    LoginRequiredMixin,
    PermissionRequiredMixin,
    CreateView,
):
    model = Adult
    form_class = AdultForm
    template_name = "adults/form.html"
    permission_required = "programs.add_adult"

    def get_initial(self):
        ini = super().get_initial()
        ini["is_mentor"] = True
        return ini

    def form_valid(self, form):
        form.instance.is_mentor = True
        return super().form_valid(form)

    def get_success_url(self):
        return reverse("mentor_list")


class MentorUpdateView(
    PassUserToFormMixin,
    SensitiveDataViewMixin,
    LogFormSaveMixin,
    LoginRequiredMixin,
    DynamicWritePermissionMixin,
    UpdateView,
):
    model = Adult
    form_class = AdultForm
    template_name = "adults/form.html"
    permission_required = "programs.change_adult"
    section = "adult_info"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["back_url"] = self.request.META.get("HTTP_REFERER", "/")
        return ctx

    def get_success_url(self):
        next_url = self.request.GET.get("next")
        safe_url = get_safe_url(self.request, next_url)
        if safe_url:
            return safe_url
        return reverse("mentor_edit", args=[self.object.pk])


# --- Schools list/create/edit ---
class SchoolListView(LoginRequiredMixin, SortableListViewMixin, ListView):
    model = School
    template_name = "schools/list.html"
    context_object_name = "schools"

    sort_fields = {
        "name": Lower("name"),
        "district": Lower("district"),
        "city": Lower("city"),
        "state": "state",
    }
    default_sort_field = "name"

    def get_queryset(self):
        return self.apply_sorting(super().get_queryset())


class SchoolCreateView(
    LogFormSaveMixin, LoginRequiredMixin, PermissionRequiredMixin, CreateView
):
    model = School
    form_class = SchoolForm
    template_name = "schools/form.html"
    permission_required = "programs.add_school"

    def get_success_url(self):
        # After creating a School, return to the Schools listing
        return reverse("school_list")


class SchoolUpdateView(
    LogFormSaveMixin, LoginRequiredMixin, PermissionRequiredMixin, UpdateView
):
    model = School
    form_class = SchoolForm
    template_name = "schools/form.html"
    permission_required = "programs.change_school"

    def get_success_url(self):
        next_url = self.request.GET.get("next")
        safe_url = get_safe_url(self.request, next_url)
        if safe_url:
            return safe_url
        return reverse("school_edit", args=[self.object.pk])


class ProgramEmailView(LoginRequiredMixin, LeadMentorRequiredMixin, View):
    template_name = "programs/email_form.html"

    def get(self, request, pk=None):
        program = get_object_or_404(Program, pk=pk) if pk else None
        form = ProgramEmailForm(program=program) if program else ProgramEmailForm()
        return self._render(form, program)

    def post(self, request, pk=None):
        program = get_object_or_404(Program, pk=pk) if pk else None
        form = (
            ProgramEmailForm(request.POST, program=program)
            if program
            else ProgramEmailForm(request.POST)
        )
        if form.is_valid():
            prog = program or form.cleaned_data["program"]
            groups = form.cleaned_data["recipient_groups"]
            subject = form.cleaned_data["subject"]
            html_body = form.cleaned_data["body"]
            # Inline CSS for better email client compatibility
            try:
                inlined_html_body = transform(html_body)
            except Exception:
                inlined_html_body = html_body
            text_body = strip_tags(inlined_html_body)
            test_email = form.cleaned_data.get("test_email")

            recipients = set()
            if "students" in groups:
                for s in Student.objects.filter(
                    enrollment__program=prog, enrollment__active=True, graduated=False
                ).distinct():
                    if s.personal_email:
                        recipients.add(s.personal_email)
                    elif s.andrew_email:
                        recipients.add(s.andrew_email)
            if "parents" in groups:
                for parent in Adult.objects.filter(
                    students__enrollment__program=prog,
                    students__enrollment__active=True,
                    email_updates=True,
                    active=True,
                ).distinct():
                    e = parent.personal_email or parent.andrew_email
                    if e:
                        recipients.add(e)
            if "mentors" in groups:
                for m in Adult.objects.filter(is_mentor=True, active=True):
                    e = m.personal_email or m.andrew_email
                    if e:
                        recipients.add(e)

            if not recipients and not test_email:
                messages.error(request, "No recipients found for the selected groups.")
                return self._render(form, prog)

            to_send = [test_email] if test_email else sorted(recipients)

            # Determine sender account and SMTP credentials
            selected = form.cleaned_data.get("from_account")
            accounts = getattr(settings, "EMAIL_SENDER_ACCOUNTS", []) or []
            acc = None
            if accounts and selected and selected != "DEFAULT":
                # Match by key or email value
                for a in accounts:
                    key = a.get("key") or a.get("email")
                    if key == selected:
                        acc = a
                        break
            # Build SMTP connection using selected account credentials if provided
            conn_kwargs = {
                "backend": getattr(
                    settings,
                    "EMAIL_BACKEND",
                    "django.core.mail.backends.smtp.EmailBackend",
                ),
                "host": getattr(settings, "EMAIL_HOST", ""),
                "port": getattr(settings, "EMAIL_PORT", 465),
                "use_tls": getattr(settings, "EMAIL_USE_TLS", False),
                "use_ssl": getattr(settings, "EMAIL_USE_SSL", True),
                "timeout": getattr(settings, "EMAIL_TIMEOUT", 10),
            }
            if acc:
                conn_kwargs.update(
                    {
                        "username": acc.get("username") or "",
                        "password": acc.get("password") or "",
                    }
                )
                from_email = acc.get("email") or getattr(
                    settings, "DEFAULT_FROM_EMAIL", "no-reply@example.com"
                )
                # Include display_name if provided
                display_name = acc.get("display_name")
                if display_name:
                    from_email = f'"{display_name}" <{from_email}>'
            else:
                # Fall back to global credentials and default from address
                conn_kwargs.update(
                    {
                        "username": getattr(settings, "EMAIL_HOST_USER", ""),
                        "password": getattr(settings, "EMAIL_HOST_PASSWORD", ""),
                    }
                )
                from_email = getattr(
                    settings, "DEFAULT_FROM_EMAIL", "no-reply@example.com"
                )
                # Include sender name from settings if available
                sender_name = getattr(settings, "DEFAULT_FROM_NAME", None)
                if sender_name:
                    from_email = f'"{sender_name}" <{from_email}>'

            connection = get_connection(**conn_kwargs)
            # For test sends, put recipient in the To field (some SMTP providers reject emails with empty To)
            if test_email:
                email = EmailMultiAlternatives(
                    subject=subject,
                    body=text_body,
                    from_email=from_email,
                    to=[test_email],
                    connection=connection,
                )
                email.bcc = []
            else:
                email = EmailMultiAlternatives(
                    subject=subject,
                    body=text_body,
                    from_email=from_email,
                    to=[],
                    connection=connection,
                )
                email.to = []  # ensure empty
                email.bcc = to_send
            email.attach_alternative(inlined_html_body, "text/html")

            # Log details about the outgoing message
            preview_recipients = to_send[:20]
            logger.info(
                "ProgramEmail: preparing to send email | from=%s | to_count=%d | subject=%s | test=%s",
                from_email,
                len(to_send),
                subject,
                bool(test_email),
            )
            logger.debug(
                "ProgramEmail: recipient sample (first %d): %s",
                len(preview_recipients),
                preview_recipients,
            )

            try:
                sent_count = email.send(fail_silently=False)
                logger.info(
                    "ProgramEmail: email sent successfully | from=%s | to_count=%d | subject=%s | sent_count=%s",
                    from_email,
                    len(to_send),
                    subject,
                    sent_count,
                )
                messages.success(
                    request,
                    f"Email sent to {len(to_send)} recipient(s){' (test only)' if test_email else ''}.",
                )
                # Redirect back to program detail if coming from there, otherwise stay
                if pk:
                    return redirect("program_detail", pk=pk)
                return redirect("program_messaging")
            except Exception as e:
                logger.error(
                    "ProgramEmail: email send FAILED | from=%s | to_count=%d | subject=%s | error=%s",
                    from_email,
                    len(to_send),
                    subject,
                    e,
                    exc_info=True,
                )
                messages.error(request, f"Failed to send email: {e}")
                return self._render(form, prog)

        return self._render(form, program)

    def _render(self, form, program):
        from django.shortcuts import render

        ctx = {"form": form, "program": program}
        return render(self.request, self.template_name, ctx)


class ProgramDetailView(LoginRequiredMixin, DynamicReadPermissionMixin, DetailView):
    model = Program
    template_name = "programs/detail.html"
    context_object_name = "program"
    section = "programs"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)

        ctx["role"] = get_user_role(self.request.user)
        program = self.object
        from .permission_views import can_user_read, can_user_write

        role = ctx["role"]

        # Prepare annotated queryset for consistent sorting
        from django.db.models import Value
        from django.db.models.functions import Coalesce, Lower, NullIf

        base_qs = (
            Enrollment.objects.filter(program=program)
            .select_related("student", "student__user", "team", "crew")
            .annotate(
                sort_first=Lower(
                    Coalesce(
                        NullIf("student__first_name", Value("")),
                        "student__legal_first_name",
                    )
                ),
                sort_last=Lower("student__last_name"),
            )
        )

        # Parent restriction
        if role == "Parent":
            try:
                adult = self.request.user.adult_profile
                base_qs = base_qs.filter(student__adults=adult)
            except (Adult.DoesNotExist, AttributeError):
                base_qs = Enrollment.objects.none()

        # Split into active and inactive sections
        ctx["active_enrollments"] = base_qs.filter(
            active=True, student__graduated=False
        ).order_by("sort_first", "sort_last")
        ctx["inactive_enrollments"] = base_qs.exclude(
            active=True, student__graduated=False
        ).order_by("sort_first", "sort_last")

        # Backwards compatibility (old templates may rely on a single list)
        ctx["active_students"] = [e.student for e in ctx["active_enrollments"]]
        ctx["inactive_students"] = [e.student for e in ctx["inactive_enrollments"]]
        ctx["enrolled_students"] = ctx["active_students"] + ctx["inactive_students"]

        ctx["teams"] = Team.objects.all()
        ctx["crews"] = program.crews.all()

        if role == "Mentor":
            ctx["can_manage_students"] = False
            ctx["can_add_payment"] = False
            ctx["can_manage_fees"] = False
            ctx["can_view_payments"] = False
            ctx["can_view_attendance"] = False
        else:
            ctx["can_manage_students"] = can_user_write(
                self.request.user, "student_info"
            )
            ctx["can_add_payment"] = can_user_write(self.request.user, "payments")
            ctx["can_manage_fees"] = can_user_write(self.request.user, "fees")
            ctx["can_view_payments"] = can_user_read(self.request.user, "payments")
            ctx["can_view_attendance"] = can_user_read(self.request.user, "attendance")

        # Document management: any user who can edit the program can manage
        # the blank documents attached to it (used by the application wizard
        # Step 9 signed-document upload flow).
        ctx["can_manage_documents"] = self.request.user.has_perm(
            "programs.change_program"
        )
        ctx["program_documents"] = program.documents.all().order_by(
            "display_order", "name"
        )

        if ctx["can_manage_students"]:
            ctx["add_existing_form"] = AddExistingStudentToProgramForm(program=program)
            ctx["quick_create_form"] = QuickCreateStudentForm()
        return ctx


class ProgramCreateView(LogFormSaveMixin, CreateView):
    model = Program
    form_class = ProgramForm
    template_name = "programs/form.html"

    def get_success_url(self):
        return reverse("program_detail", args=[self.object.pk])


class ProgramUpdateView(LogFormSaveMixin, UpdateView):
    model = Program
    form_class = ProgramForm
    template_name = "programs/form.html"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        program = self.object
        ctx["can_manage_documents"] = self.request.user.has_perm(
            "programs.change_program"
        )
        ctx["program_documents"] = program.documents.all().order_by(
            "display_order", "name"
        )
        return ctx

    def get_success_url(self):
        return reverse("program_detail", args=[self.object.pk])


# --- Student edit ---
class StudentUpdateView(
    PassUserToFormMixin,
    SensitiveDataViewMixin,
    LogFormSaveMixin,
    LoginRequiredMixin,
    DynamicWritePermissionMixin,
    UpdateView,
):
    model = Student
    form_class = StudentForm
    template_name = "students/form.html"
    permission_required = "programs.change_student"
    section = "student_info"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)

        ctx["role"] = get_user_role(self.request.user)
        ctx["RELATIONSHIP_CHOICES"] = RELATIONSHIP_CHOICES
        student = self.object
        # Union of enabled feature keys across all enrolled programs
        keys = (
            set(
                k
                for k in student.programs.values_list(
                    "features__key", flat=True
                ).distinct()
                if k
            )
            if student
            else set()
        )
        ctx["program_feature_keys"] = keys
        if student:
            ctx["enrollments"] = student.enrollment_set.all().select_related("program")
        return ctx

    def form_valid(self, form):
        response = super().form_valid(form)
        # Persist relationship selections for each selected parent
        rel_map = {
            k[len("parent_rel_") :]: v  # noqa: E203
            for k, v in self.request.POST.items()
            if k.startswith("parent_rel_")
        }
        valid_keys = set(k for k, _ in RELATIONSHIP_CHOICES)
        for pid_str, rel in rel_map.items():
            try:
                pid = int(pid_str)
            except (TypeError, ValueError):
                continue
            if rel in valid_keys:
                AdultStudentRelationship.objects.update_or_create(
                    adult_id=pid,
                    student=self.object,
                    defaults={"relationship_to_student": rel},
                )
        messages.success(self.request, "Student record saved successfully.")
        return response

    def get_success_url(self):
        next_url = self.request.GET.get("next")
        safe_url = get_safe_url(self.request, next_url)
        if safe_url:
            return safe_url
        return reverse("student_detail", args=[self.object.pk])


class StudentCreateView(
    PassUserToFormMixin,
    LogFormSaveMixin,
    LoginRequiredMixin,
    PermissionRequiredMixin,
    DynamicWritePermissionMixin,
    CreateView,
):
    model = Student
    form_class = StudentForm
    template_name = "students/form.html"
    permission_required = "programs.add_student"
    section = "student_info"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)

        ctx["role"] = get_user_role(self.request.user)
        ctx["RELATIONSHIP_CHOICES"] = RELATIONSHIP_CHOICES
        return ctx

    def form_valid(self, form):
        response = super().form_valid(form)
        # Persist relationship selections for each selected parent
        rel_map = {
            k[len("parent_rel_") :]: v  # noqa: E203
            for k, v in self.request.POST.items()
            if k.startswith("parent_rel_")
        }
        valid_keys = set(k for k, _ in RELATIONSHIP_CHOICES)
        for pid_str, rel in rel_map.items():
            try:
                pid = int(pid_str)
            except (TypeError, ValueError):
                continue
            if rel in valid_keys:
                AdultStudentRelationship.objects.update_or_create(
                    adult_id=pid,
                    student=self.object,
                    defaults={"relationship_to_student": rel},
                )
        return response

    def get_success_url(self):
        # After creating a Student, return to the Students listing
        return reverse("student_list")


class AdultDetailView(
    DynamicPermissionMixin, SensitiveDataViewMixin, LoginRequiredMixin, DetailView
):
    model = Adult
    template_name = "adults/detail.html"
    context_object_name = "adult"
    section = "adult_info"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)

        ctx["role"] = get_user_role(self.request.user)
        return ctx


class StudentDetailView(
    DynamicPermissionMixin, SensitiveDataViewMixin, LoginRequiredMixin, DetailView
):
    model = Student
    template_name = "students/detail.html"
    context_object_name = "student"
    section = "student_info"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)

        ctx["role"] = get_user_role(self.request.user)
        student = self.object
        # Union of enabled feature keys across all enrolled programs
        keys = set(
            k
            for k in student.programs.values_list("features__key", flat=True).distinct()
            if k
        )
        ctx["program_feature_keys"] = keys
        return ctx


# --- Program student management actions ---
class ProgramStudentAddView(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = "programs.change_student"

    def post(self, request, pk):
        program = get_object_or_404(Program, pk=pk)
        form = AddExistingStudentToProgramForm(request.POST, program=program)
        if form.is_valid():
            student = form.cleaned_data["student"]
            Enrollment.objects.get_or_create(student=student, program=program)
            messages.success(request, f"Added {student} to {program}.")
        else:
            messages.error(request, "Could not add student to program.")
        return redirect("program_detail", pk=program.pk)


class ProgramStudentQuickCreateView(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = "programs.add_student"

    def post(self, request, pk):
        program = get_object_or_404(Program, pk=pk)
        form = QuickCreateStudentForm(request.POST)
        if form.is_valid():
            student = form.save()
            Enrollment.objects.get_or_create(student=student, program=program)
            messages.success(request, f"Created {student} and added to {program}.")
        else:
            messages.error(request, "Could not create student.")
        return redirect("program_detail", pk=program.pk)


class ProgramEnrollmentUpdateView(LoginRequiredMixin, LeadMentorRequiredMixin, View):
    def post(self, request, pk):
        enrollment_id = request.POST.get("enrollment_id")
        team_id = request.POST.get("team_id")
        crew_id = request.POST.get("crew_id")
        subteam_id = request.POST.get("subteam_id")
        active = request.POST.get("active")
        enrollment = get_object_or_404(Enrollment, id=enrollment_id, program_id=pk)

        updated_fields = []
        if active is not None:
            new_active = active.lower() == "true"
            if enrollment.active != new_active:
                enrollment.active = new_active
                updated_fields.append("Active status")

        if team_id is not None:
            if team_id:
                team = get_object_or_404(Team, id=team_id)
                enrollment.team = team
            else:
                enrollment.team = None
            updated_fields.append("Team")

        if crew_id is not None:
            if crew_id:
                crew = get_object_or_404(Crew, id=crew_id, program_id=pk)
                enrollment.crew = crew
            else:
                enrollment.crew = None
            updated_fields.append("Crew")

        if subteam_id is not None:
            if subteam_id:
                subteam = get_object_or_404(SubTeam, id=subteam_id, program_id=pk)
                enrollment.subteam = subteam
            else:
                enrollment.subteam = None
            updated_fields.append("SubTeam")

        enrollment.save()
        if updated_fields:
            messages.success(
                request,
                f"{' and '.join(updated_fields)} updated for {enrollment.student}.",
            )
        next_url = request.POST.get("next")
        safe_url = get_safe_url(request, next_url)
        if safe_url:
            return redirect(safe_url)
        return redirect("program_detail", pk=pk)


class ProgramStudentRemoveView(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = "programs.change_student"

    def post(self, request, pk, student_id):
        program = get_object_or_404(Program, pk=pk)
        student = get_object_or_404(Student, pk=student_id)
        Enrollment.objects.filter(student=student, program=program).delete()
        messages.success(request, f"Removed {student} from {program}.")
        return redirect("program_detail", pk=program.pk)


class ProgramAssignmentView(LoginRequiredMixin, LeadMentorRequiredMixin, View):
    template_name = "programs/assignment.html"

    def get(self, request, pk):
        program = get_object_or_404(Program, pk=pk)
        enrollments = Enrollment.objects.filter(program=program).select_related(
            "student", "team", "crew", "subteam"
        )
        teams = Team.objects.all().order_by("number")
        crews = Crew.objects.filter(program=program).order_by("name")
        subteams = SubTeam.objects.filter(program=program).order_by("name")

        # Separate inactive students (inactive enrollment or graduated) so they
        # don't get mixed in with the active ones being assigned.
        active_enrollments = enrollments.filter(active=True, student__graduated=False)
        inactive_enrollments = enrollments.exclude(
            active=True, student__graduated=False
        )

        return render(
            request,
            self.template_name,
            {
                "program": program,
                "enrollments": enrollments,
                "active_enrollments": active_enrollments,
                "inactive_enrollments": inactive_enrollments,
                "teams": teams,
                "crews": crews,
                "subteams": subteams,
            },
        )

    def post(self, request, pk):
        program = get_object_or_404(Program, pk=pk)
        assignment_type = request.POST.get("assignment_type")
        target_id = request.POST.get("target_id")
        student_ids = request.POST.getlist("student_ids")

        if not student_ids:
            messages.warning(request, "No students selected.")
            return redirect("program_assignment", pk=pk)

        if not target_id:
            messages.warning(request, f"No {assignment_type} selected.")
            return redirect("program_assignment", pk=pk)

        enrollments = Enrollment.objects.filter(
            program=program, student_id__in=student_ids
        )

        if assignment_type == "team":
            team = get_object_or_404(Team, id=target_id)
            enrollments.update(team=team)
            messages.success(
                request, f"Assigned {len(student_ids)} students to team {team}."
            )
        elif assignment_type == "crew":
            crew = get_object_or_404(Crew, id=target_id, program=program)
            enrollments.update(crew=crew)
            messages.success(
                request, f"Assigned {len(student_ids)} students to crew {crew}."
            )
        elif assignment_type == "subteam":
            subteam = get_object_or_404(SubTeam, id=target_id, program=program)
            enrollments.update(subteam=subteam)
            messages.success(
                request, f"Assigned {len(student_ids)} students to subteam {subteam}."
            )

        return redirect("program_assignment", pk=pk)


# --- Parent create/edit ---
class ParentCreateView(
    PassUserToFormMixin,
    LogFormSaveMixin,
    LoginRequiredMixin,
    PermissionRequiredMixin,
    CreateView,
):
    model = Adult
    form_class = AdultForm
    template_name = "adults/form.html"
    permission_required = "programs.add_adult"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["back_url"] = self.request.META.get("HTTP_REFERER", "/")
        return ctx

    def form_valid(self, form):
        # Ensure adults created via this view are flagged as parents
        obj = form.save(commit=False)
        obj.is_parent = True
        obj.save()
        # Save many-to-many after the object exists
        form.save_m2m()
        # Logging for creation with changed fields
        user = getattr(self.request, "user", None)
        user_repr = (
            f"{getattr(user, 'pk', 'anon')}:{getattr(user, 'username', 'anonymous')}"
            if getattr(user, "is_authenticated", False)
            else "anonymous"
        )
        for f in getattr(form, "changed_data", []) or []:
            new = form.cleaned_data.get(f, getattr(obj, f, None))
            forms_logger.info(
                "FormSave: %s[%s] %s by %s | field=%s | from=%s | to=%s",
                "Adult",
                obj.pk,
                "create",
                user_repr,
                f,
                self._fmt_val(None),
                self._fmt_val(new),
            )
        messages.success(self.request, "Parent added successfully.")
        return redirect("parent_list")

    def get_success_url(self):
        # After creating a Parent, return to the Parents listing
        return reverse("parent_list")


class ParentUpdateView(
    PassUserToFormMixin,
    SensitiveDataViewMixin,
    LogFormSaveMixin,
    LoginRequiredMixin,
    DynamicWritePermissionMixin,
    UpdateView,
):
    model = Adult
    form_class = AdultForm
    template_name = "adults/form.html"
    permission_required = "programs.change_adult"
    section = "adult_info"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["back_url"] = self.request.META.get("HTTP_REFERER", "/")
        return ctx

    def get_success_url(self):
        next_url = self.request.GET.get("next")
        safe_url = get_safe_url(self.request, next_url)
        if safe_url:
            return safe_url
        return reverse("parent_edit", args=[self.object.pk])


# --- Payment create ---
class ProgramPaymentCreateView(
    LogFormSaveMixin,
    LoginRequiredMixin,
    DynamicWritePermissionMixin,
    CreateView,
):
    model = Payment
    form_class = PaymentForm
    template_name = "programs/payment_form.html"
    section = "payments"

    def dispatch(self, request, *args, **kwargs):
        self.program = get_object_or_404(Program, pk=kwargs["pk"])
        return super().dispatch(request, *args, **kwargs)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["program"] = self.program
        return kwargs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["program"] = self.program
        return ctx

    def form_valid(self, form):
        obj = form.save(commit=False)
        # Ensure program is set from the URL context
        obj.program = self.program
        obj.save()
        # Log creation with a concise summary and field values
        user = getattr(self.request, "user", None)
        user_repr = (
            f"{getattr(user, 'pk', 'anon')}:{getattr(user, 'username', 'anonymous')}"
            if getattr(user, "is_authenticated", False)
            else "anonymous"
        )
        forms_logger.info(
            "FormSave: Payment[%s] create by %s | student=%s | program=%s | amount=%s | paid_via=%s | paid_on=%s",
            obj.pk,
            user_repr,
            self._fmt_val(getattr(obj, "student", None)),
            self._fmt_val(getattr(obj, "program", None)),
            self._fmt_val(getattr(obj, "amount", None)),
            self._fmt_val(getattr(obj, "paid_via", None)),
            self._fmt_val(getattr(obj, "paid_on", None)),
        )
        messages.success(self.request, "Payment recorded successfully.")
        return redirect("program_detail", pk=self.program.pk)


class ProgramPaymentDetailView(LoginRequiredMixin, DynamicReadPermissionMixin, View):
    section = "payments"

    def get_object(self):
        return get_object_or_404(Payment, pk=self.kwargs["payment_id"])

    def get(self, request, pk, payment_id):
        program = get_object_or_404(Program, pk=pk)
        payment = get_object_or_404(Payment, pk=payment_id)
        # Ensure payment belongs to this program
        if payment.program_id != program.id:
            messages.error(request, "Payment does not belong to this program.")
            return redirect("program_detail", pk=program.pk)
        student = payment.student
        # Ensure enrollment
        if not Enrollment.objects.filter(student=student, program=program).exists():
            messages.error(request, f"{student} is not enrolled in {program}.")
            return redirect("program_detail", pk=program.pk)
        from django.shortcuts import render

        return render(
            request,
            "programs/payment_detail.html",
            {
                "program": program,
                "student": student,
                "payment": payment,
            },
        )


class ProgramPaymentPrintView(LoginRequiredMixin, DynamicReadPermissionMixin, View):
    section = "payments"

    def get_object(self):
        return get_object_or_404(Payment, pk=self.kwargs["payment_id"])

    def get(self, request, pk, payment_id):
        program = get_object_or_404(Program, pk=pk)
        payment = get_object_or_404(Payment, pk=payment_id)
        if payment.program_id != program.id:
            messages.error(request, "Payment does not belong to this program.")
            return redirect("program_detail", pk=program.pk)
        student = payment.student
        if not Enrollment.objects.filter(student=student, program=program).exists():
            messages.error(request, f"{student} is not enrolled in {program}.")
            return redirect("program_detail", pk=program.pk)
        from django.shortcuts import render

        return render(
            request,
            "programs/payment_print.html",
            {
                "program": program,
                "student": student,
                "payment": payment,
            },
        )


class SlidingScaleCreateView(
    LogFormSaveMixin,
    LoginRequiredMixin,
    DynamicWritePermissionMixin,
    CreateView,
):
    model = SlidingScale
    form_class = SlidingScaleForm
    template_name = "programs/sliding_scale_form.html"
    section = "sliding_scale"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["settings_obj"] = SlidingScaleSettings.get_solo()
        return ctx

    def form_valid(self, form):
        # The sliding scale is no longer tied to a single program — it applies
        # across all of the student's programs — so we simply create the record.
        obj = form.save(commit=False)
        obj.save()
        # Log creation
        user = getattr(self.request, "user", None)
        user_repr = (
            f"{getattr(user, 'pk', 'anon')}:{getattr(user, 'username', 'anonymous')}"
            if getattr(user, "is_authenticated", False)
            else "anonymous"
        )
        forms_logger.info(
            "FormSave: SlidingScale[%s] create by %s | student=%s | percent=%s",
            obj.pk,
            user_repr,
            self._fmt_val(getattr(obj, "student", None)),
            self._fmt_val(getattr(obj, "percent", None)),
        )
        messages.success(self.request, "Sliding scale saved successfully.")
        return redirect("sliding_scale_review_list")


class SlidingScaleUpdateView(
    LogFormSaveMixin,
    LoginRequiredMixin,
    DynamicWritePermissionMixin,
    UpdateView,
):
    model = SlidingScale
    form_class = SlidingScaleForm
    template_name = "programs/sliding_scale_form.html"
    section = "sliding_scale"

    def get_object(self, queryset=None):
        return get_object_or_404(SlidingScale, pk=self.kwargs["sliding_id"])

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["settings_obj"] = SlidingScaleSettings.get_solo()
        return ctx

    def form_valid(self, form):
        obj = form.save(commit=False)
        # Capture old values for changed fields before saving
        try:
            before = SlidingScale.objects.get(pk=obj.pk)
        except SlidingScale.DoesNotExist:
            before = None
        obj.save()
        # Log update with field-level changes when possible
        user = getattr(self.request, "user", None)
        user_repr = (
            f"{getattr(user, 'pk', 'anon')}:{getattr(user, 'username', 'anonymous')}"
            if getattr(user, "is_authenticated", False)
            else "anonymous"
        )
        for f in getattr(form, "changed_data", []) or []:
            old = getattr(before, f, None) if before is not None else None
            new = getattr(obj, f, None)
            forms_logger.info(
                "FormSave: %s[%s] %s by %s | field=%s | from=%s | to=%s",
                "SlidingScale",
                obj.pk,
                "update",
                user_repr,
                f,
                self._fmt_val(old),
                self._fmt_val(new),
            )
        messages.success(self.request, "Sliding scale updated successfully.")
        return redirect("sliding_scale_review_list")


class SlidingScaleTaxFormDeleteView(
    LoginRequiredMixin, DynamicWritePermissionMixin, View
):
    section = "sliding_scale"
    permission_required = "programs.change_slidingscale"

    def test_func(self):
        # Allow users with change_slidingscale permission in addition to LeadMentors
        if self.request.user.has_perm("programs.change_slidingscale"):
            return True
        return super().test_func()

    def post(self, request, sliding_id, form_id):
        tax_form = get_object_or_404(
            TaxForm,
            pk=form_id,
            sliding_scale_id=sliding_id,
        )
        tax_form.file.delete(save=False)
        tax_form.delete()
        messages.success(request, "Tax form deleted.")
        return redirect("sliding_scale_edit", sliding_id=sliding_id)


class SlidingScaleTaxFormViewView(LoginRequiredMixin, LeadMentorRequiredMixin, View):
    """Stream the *decrypted* contents of an uploaded tax form so a Lead
    Mentor can view it (e.g. a PDF or image rendered inline in the browser)
    or download it, without ever exposing the encrypted bytes on disk."""

    def test_func(self):
        # Allow users with change_slidingscale permission in addition to LeadMentors
        if self.request.user.has_perm("programs.change_slidingscale"):
            return True
        return super().test_func()

    def get(self, request, sliding_id, form_id):
        tax_form = get_object_or_404(TaxForm, pk=form_id, sliding_scale_id=sliding_id)
        filename = os.path.basename(tax_form.file.name)
        content_type, _ = mimetypes.guess_type(filename)
        content_type = content_type or "application/octet-stream"
        disposition = "attachment" if request.GET.get("download") else "inline"
        response = FileResponse(
            tax_form.file.open("rb"),
            content_type=content_type,
        )
        response["Content-Disposition"] = f'{disposition}; filename="{filename}"'
        return response


class SlidingScaleReviewListView(LoginRequiredMixin, LeadMentorRequiredMixin, ListView):
    """Lead Mentor queue of sliding scale applications awaiting review."""

    model = SlidingScale
    template_name = "programs/sliding_scale_review_list.html"
    context_object_name = "applications"

    def get_queryset(self):
        return (
            SlidingScale.objects.filter(status=SlidingScale.STATUS_PENDING)
            .select_related("student", "applied_by")
            .order_by("created_at")
        )

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["decided_applications"] = (
            SlidingScale.objects.exclude(status=SlidingScale.STATUS_PENDING)
            .select_related("student", "applied_by", "reviewed_by")
            .order_by("-reviewed_at", "-updated_at")[:25]
        )
        return ctx


class SlidingScaleReviewDecideView(LoginRequiredMixin, LeadMentorRequiredMixin, View):
    """Approve or decline a pending sliding scale application."""

    template_name = "programs/sliding_scale_review_detail.html"

    def get_object(self):
        return get_object_or_404(SlidingScale, pk=self.kwargs["pk"])

    def get(self, request, pk):
        application = self.get_object()
        settings_obj = SlidingScaleSettings.get_solo()
        suggested_percent = settings_obj.compute_discount_percent(
            application.family_size, application.adjusted_gross_income
        )
        return render(
            request,
            self.template_name,
            {
                "application": application,
                "suggested_percent": suggested_percent,
                "settings_obj": settings_obj,
            },
        )

    def post(self, request, pk):
        application = self.get_object()
        action = request.POST.get("action")

        if action == "approve":
            percent = request.POST.get("percent")
            date_val = request.POST.get("date") or None
            expiration_date = request.POST.get("expiration_date") or None
            try:
                application.percent = Decimal(percent)
            except (TypeError, InvalidOperation):
                messages.error(request, "Please enter a valid discount percent.")
                return redirect("sliding_scale_review_decide", pk=pk)
            if application.percent < 0 or application.percent > 100:
                messages.error(request, "Percent must be between 0 and 100.")
                return redirect("sliding_scale_review_decide", pk=pk)
            application.date = date_val or datetime.date.today()
            application.expiration_date = expiration_date
            application.status = SlidingScale.STATUS_APPROVED
            application.reviewed_by = request.user
            application.reviewed_at = timezone.now()
            application.save()
            messages.success(
                request, f"Sliding scale approved for {application.student}."
            )
        elif action == "decline":
            reason = (request.POST.get("decline_reason") or "").strip()
            if not reason:
                messages.error(
                    request, "Please provide a reason for declining this application."
                )
                return redirect("sliding_scale_review_decide", pk=pk)
            application.status = SlidingScale.STATUS_DECLINED
            application.decline_reason = reason
            application.reviewed_by = request.user
            application.reviewed_at = timezone.now()
            application.save()
            messages.success(
                request,
                f"Sliding scale application declined for {application.student}.",
            )
        else:
            messages.error(request, "Unknown action.")

        return redirect("sliding_scale_review_list")


class SlidingScaleApplyView(LoginRequiredMixin, View):
    """Parent-facing sliding scale application, reached from the Payments page.

    Only a Parent (not the Student, and not a Mentor) may apply, and only for
    their own linked student(s). The application applies across all of the
    student's programs, not just one.
    """

    template_name = "programs/sliding_scale_apply.html"

    def _get_student_for_parent(self, request, student_id):
        if get_user_role(request.user) != "Parent":
            messages.error(request, "Only a parent can apply for the sliding scale.")
            return None, redirect("parent_payments")

        student = get_object_or_404(Student, pk=student_id)
        try:
            adult = request.user.adult_profile
            if not adult.is_parent or student not in adult.students.all():
                raise Adult.DoesNotExist
        except (Adult.DoesNotExist, AttributeError):
            messages.error(
                request, "You do not have permission to apply for this student."
            )
            return None, redirect("parent_payments")

        return student, None

    def get(self, request, student_id):
        student, error_redirect = self._get_student_for_parent(request, student_id)
        if error_redirect:
            return error_redirect

        existing_pending = SlidingScale.objects.filter(
            student=student, status=SlidingScale.STATUS_PENDING
        ).exists()
        if existing_pending:
            messages.info(
                request,
                f"{student} already has a sliding scale application pending review.",
            )
            return redirect("parent_payments")

        form = SlidingScaleApplicationForm()
        settings_obj = SlidingScaleSettings.get_solo()
        return render(
            request,
            self.template_name,
            {"student": student, "form": form, "settings_obj": settings_obj},
        )

    def post(self, request, student_id):
        student, error_redirect = self._get_student_for_parent(request, student_id)
        if error_redirect:
            return error_redirect

        form = SlidingScaleApplicationForm(request.POST, request.FILES)
        if not form.is_valid():
            settings_obj = SlidingScaleSettings.get_solo()
            return render(
                request,
                self.template_name,
                {"student": student, "form": form, "settings_obj": settings_obj},
            )

        adult = request.user.adult_profile
        application = SlidingScale.objects.create(
            student=student,
            family_size=form.cleaned_data["family_size"],
            adjusted_gross_income=form.cleaned_data["adjusted_gross_income"],
            status=SlidingScale.STATUS_PENDING,
            applied_by=adult,
            notes=form.cleaned_data.get("notes") or "",
        )

        documents = form.cleaned_data.get("documents")
        if documents:
            files = documents if isinstance(documents, list) else [documents]
            for f in files:
                TaxForm.objects.create(sliding_scale=application, file=f)

        messages.success(
            request,
            f"Sliding scale application submitted for {student}. A Lead Mentor will review it soon.",
        )
        return redirect("parent_payments")


class SlidingScaleWithdrawView(LoginRequiredMixin, View):
    """Allows a Parent to withdraw their own pending sliding scale application
    (e.g. to correct a mistake and reapply)."""

    def post(self, request, pk):
        application = get_object_or_404(
            SlidingScale, pk=pk, status=SlidingScale.STATUS_PENDING
        )
        student = application.student

        if get_user_role(request.user) != "Parent":
            messages.error(
                request, "Only a parent can withdraw a sliding scale application."
            )
            return redirect("parent_payments")

        try:
            adult = request.user.adult_profile
            if not adult.is_parent or student not in adult.students.all():
                raise Adult.DoesNotExist
        except (Adult.DoesNotExist, AttributeError):
            messages.error(
                request,
                "You do not have permission to withdraw this application.",
            )
            return redirect("parent_payments")

        for tax_form in application.tax_forms.all():
            tax_form.file.delete(save=False)
            tax_form.delete()
        application.delete()

        messages.success(
            request,
            f"Sliding scale application for {student} has been withdrawn. You may submit a new one at any time.",
        )
        return redirect("parent_payments")


class ProgramStudentBalanceView(LoginRequiredMixin, DynamicReadPermissionMixin, View):
    section = "payments"

    def get(self, request, pk, student_id):
        program = get_object_or_404(Program, pk=pk)
        student = get_object_or_404(Student, pk=student_id)

        # Object level check for Parents

        if get_user_role(request.user) == "Parent":
            try:
                adult = request.user.adult_profile
                if student not in adult.students.all():
                    messages.error(
                        request,
                        "You do not have permission to view this balance sheet.",
                    )
                    return redirect("home")
            except Exception:
                messages.error(
                    request, "You do not have permission to view this balance sheet."
                )
                return redirect("home")
        # Ensure enrollment
        if not Enrollment.objects.filter(student=student, program=program).exists():
            messages.error(request, f"{student} is not enrolled in {program}.")
            return redirect("program_detail", pk=program.pk)

        from .permission_views import can_user_read

        can_view_sliding = can_user_read(request.user, "sliding_scale")
        balance_data = get_student_balance_data(
            student, program, can_view_sliding=can_view_sliding
        )

        from django.shortcuts import render

        return render(
            request,
            "programs/balance_sheet.html",
            {
                "program": program,
                "student": student,
                "entries": balance_data["entries"],
                "total_fees": balance_data["total_fees"],
                "total_sliding": balance_data["total_sliding"],
                "total_payments": balance_data["total_payments"],
                "balance": balance_data["balance"],
                "sliding_scale": balance_data["sliding_scale"],
            },
        )


class ProgramStudentBalancePrintView(
    LoginRequiredMixin, DynamicReadPermissionMixin, View
):
    section = "payments"

    def get_object(self):
        return get_object_or_404(Student, pk=self.kwargs["student_id"])

    def get(self, request, pk, student_id):
        program = get_object_or_404(Program, pk=pk)
        student = get_object_or_404(Student, pk=student_id)

        # Object level check for Parents

        if get_user_role(request.user) == "Parent":
            try:
                adult = request.user.adult_profile
                if student not in adult.students.all():
                    messages.error(
                        request,
                        "You do not have permission to view this balance sheet.",
                    )
                    return redirect("home")
            except Exception:
                messages.error(
                    request, "You do not have permission to view this balance sheet."
                )
                return redirect("home")
        # Ensure enrollment
        if not Enrollment.objects.filter(student=student, program=program).exists():
            messages.error(request, f"{student} is not enrolled in {program}.")
            return redirect("program_detail", pk=program.pk)

        from .permission_views import can_user_read

        can_view_sliding = can_user_read(request.user, "sliding_scale")
        balance_data = get_student_balance_data(
            student, program, can_view_sliding=can_view_sliding
        )

        from django.shortcuts import render

        return render(
            request,
            "programs/balance_sheet_print.html",
            {
                "program": program,
                "student": student,
                "entries": balance_data["entries"],
                "total_fees": balance_data["total_fees"],
                "total_sliding": balance_data["total_sliding"],
                "total_payments": balance_data["total_payments"],
                "balance": balance_data["balance"],
                "sliding_scale": balance_data["sliding_scale"],
            },
        )


class ProgramFeeSelectView(LoginRequiredMixin, LeadMentorRequiredMixin, View):
    template_name = "programs/fee_select.html"

    def dispatch(self, request, *args, **kwargs):
        self.program = get_object_or_404(Program, pk=kwargs["pk"])
        return super().dispatch(request, *args, **kwargs)

    def get(self, request, pk):
        from django.shortcuts import render

        fees = Fee.objects.filter(program=self.program).order_by("name")
        return render(
            request, self.template_name, {"program": self.program, "fees": fees}
        )


class ProgramFeeAssignmentEditView(
    LoginRequiredMixin, DynamicWritePermissionMixin, View
):
    template_name = "programs/fee_assignment_form.html"
    section = "fees"

    def dispatch(self, request, *args, **kwargs):
        self.program = get_object_or_404(Program, pk=kwargs["pk"])
        self.fee = get_object_or_404(Fee, pk=kwargs["fee_id"], program=self.program)
        return super().dispatch(request, *args, **kwargs)

    def get(self, request, pk, fee_id):
        form = FeeAssignmentEditForm(program=self.program, fee=self.fee)
        from django.shortcuts import render

        return render(
            request,
            self.template_name,
            {"program": self.program, "fee": self.fee, "form": form},
        )

    def post(self, request, pk, fee_id):
        form = FeeAssignmentEditForm(request.POST, program=self.program, fee=self.fee)
        if form.is_valid():
            form.save()
            messages.success(request, "Fee applicability saved.")
            return redirect(
                "program_fee_assignments", pk=self.program.pk, fee_id=self.fee.pk
            )
        from django.shortcuts import render

        return render(
            request,
            self.template_name,
            {"program": self.program, "fee": self.fee, "form": form},
        )


class ProgramFeeCreateView(
    LogFormSaveMixin,
    LoginRequiredMixin,
    PermissionRequiredMixin,
    DynamicWritePermissionMixin,
    CreateView,
):
    permission_required = "programs.add_fee"
    model = Fee
    form_class = FeeForm
    template_name = "programs/fee_form.html"
    section = "fees"

    def dispatch(self, request, *args, **kwargs):
        self.program = get_object_or_404(Program, pk=kwargs["pk"])
        return super().dispatch(request, *args, **kwargs)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["program"] = self.program
        return kwargs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)

        ctx["role"] = get_user_role(self.request.user)
        ctx["program"] = self.program
        ctx["is_create"] = True
        return ctx

    def form_valid(self, form):
        messages.success(self.request, "Fee created.")
        return super().form_valid(form)

    def get_success_url(self):
        return reverse(
            "program_fee_assignments",
            kwargs={"pk": self.program.pk, "fee_id": self.object.pk},
        )


class ProgramFeeUpdateView(
    LogFormSaveMixin,
    LoginRequiredMixin,
    PermissionRequiredMixin,
    DynamicWritePermissionMixin,
    UpdateView,
):
    permission_required = "programs.change_fee"
    model = Fee
    form_class = FeeForm
    template_name = "programs/fee_form.html"
    pk_url_kwarg = "fee_id"
    section = "fees"

    def dispatch(self, request, *args, **kwargs):
        self.program = get_object_or_404(Program, pk=kwargs["pk"])
        return super().dispatch(request, *args, **kwargs)

    def get_object(self, queryset=None):
        return get_object_or_404(Fee, pk=self.kwargs["fee_id"], program=self.program)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["program"] = self.program
        return kwargs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["program"] = self.program
        ctx["is_create"] = False
        return ctx

    def form_valid(self, form):
        messages.success(self.request, "Fee updated.")
        return super().form_valid(form)

    def get_success_url(self):
        return reverse(
            "program_fee_assignments",
            kwargs={"pk": self.program.pk, "fee_id": self.object.pk},
        )


class ProgramEmailBalancesView(LoginRequiredMixin, LeadMentorRequiredMixin, View):
    template_name = "programs/email_balances_form.html"

    def get(self, request, pk):
        program = get_object_or_404(Program, pk=pk)
        form = ProgramEmailBalancesForm(program=program)
        return self._render(request, form, program)

    def post(self, request, pk):
        program = get_object_or_404(Program, pk=pk)
        form = ProgramEmailBalancesForm(request.POST, program=program)
        if not form.is_valid():
            return self._render(request, form, program)

        subject = form.cleaned_data["subject"]
        default_message = form.cleaned_data.get("default_message") or ""
        recipient_filter = form.cleaned_data.get("recipient_filter")
        selected_student = form.cleaned_data.get("student")
        test_email = form.cleaned_data.get("test_email")

        # Build sender connection (reuse logic from ProgramEmailView)
        selected = form.cleaned_data.get("from_account")
        accounts = getattr(settings, "EMAIL_SENDER_ACCOUNTS", []) or []
        acc = None
        if accounts and selected and selected != "DEFAULT":
            for a in accounts:
                key = a.get("key") or a.get("email")
                if key == selected:
                    acc = a
                    break
        conn_kwargs = {
            "backend": getattr(
                settings, "EMAIL_BACKEND", "django.core.mail.backends.smtp.EmailBackend"
            ),
            "host": getattr(settings, "EMAIL_HOST", ""),
            "port": getattr(settings, "EMAIL_PORT", 465),
            "use_tls": getattr(settings, "EMAIL_USE_TLS", False),
            "use_ssl": getattr(settings, "EMAIL_USE_SSL", True),
            "timeout": getattr(settings, "EMAIL_TIMEOUT", 10),
        }
        if acc:
            conn_kwargs.update(
                {
                    "username": acc.get("username") or "",
                    "password": acc.get("password") or "",
                }
            )
            from_email = acc.get("email") or getattr(
                settings, "DEFAULT_FROM_EMAIL", "no-reply@example.com"
            )
            # Include display_name name if provided
            display_name = acc.get("display_name")
            if display_name:
                from_email = f'"{display_name}" <{from_email}>'
        else:
            conn_kwargs.update(
                {
                    "username": getattr(settings, "EMAIL_HOST_USER", ""),
                    "password": getattr(settings, "EMAIL_HOST_PASSWORD", ""),
                }
            )
            from_email = getattr(settings, "DEFAULT_FROM_EMAIL", "no-reply@example.com")
            # Include sender name from settings if available
            sender_name = getattr(settings, "DEFAULT_FROM_NAME", None)
            if sender_name:
                from_email = f'"{sender_name}" <{from_email}>'
        connection = get_connection(**conn_kwargs)

        # Collect students enrolled in program
        students_qs = Student.objects.filter(
            enrollment__program=program
        ).select_related("school")
        if recipient_filter == "individual" and selected_student:
            students_qs = students_qs.filter(pk=selected_student.pk)

        students = students_qs.order_by(
            Lower(Coalesce(NullIf("first_name", Value("")), "legal_first_name")),
            Lower("last_name"),
        )

        from .permission_views import can_user_read

        can_view_sliding = can_user_read(self.request.user, "sliding_scale")

        # Build list of targets with non-empty recipient emails
        targets = []
        for s in students:
            balance_data = get_student_balance_data(
                s, program, can_view_sliding=can_view_sliding
            )
            entries = balance_data["entries"]
            total_fees = balance_data["total_fees"]
            total_sliding = balance_data["total_sliding"]
            total_payments = balance_data["total_payments"]
            balance = balance_data["balance"]
            sliding = balance_data["sliding_scale"]

            # Apply recipient filters
            if recipient_filter == "non_zero" and balance == 0:
                continue
            if recipient_filter == "positive" and balance <= 0:
                continue

            # Gather recipient emails: only parents/guardians who opted in for updates
            emails = []
            for adult in s.all_parents:
                # Only include parents/guardians who have opted into email updates and are active
                if getattr(adult, "email_updates", False) and getattr(
                    adult, "active", True
                ):
                    email = adult.personal_email or adult.andrew_email
                    if email:
                        emails.append(email)
            # Deduplicate while preserving order
            seen = set()
            deduped = []
            for e in emails:
                if e and e not in seen:
                    deduped.append(e)
                    seen.add(e)
            if not deduped:
                continue
            targets.append(
                {
                    "student": s,
                    "emails": deduped,
                    "entries": entries,
                    "total_fees": total_fees,
                    "total_sliding": total_sliding,
                    "total_payments": total_payments,
                    "balance": balance,
                    "sliding_scale": sliding,
                }
            )

        if not targets and not test_email:
            messages.error(request, "No recipients found to email.")
            return self._render(request, form, program)

        # Prepare sending: if test, pick first student's content or a generic minimal body
        to_send = []
        if test_email:
            sample = targets[0] if targets else None
            if sample is None:
                messages.error(
                    request, "No sample data available to send a test email."
                )
                return self._render(request, form, program)
            to_send.append((test_email, sample))
        else:
            for t in targets:
                # send one email to combined recipients per student
                to_send.append((t["emails"], t))

        sent_total = 0
        for dest, data in to_send:
            # Render balance sheet HTML
            ctx = {
                "program": program,
                "student": data["student"],
                "entries": data["entries"],
                "total_fees": data["total_fees"],
                "total_sliding": data["total_sliding"],
                "total_payments": data["total_payments"],
                "balance": data["balance"],
                "sliding_scale": data["sliding_scale"],
            }
            # Include optional rich-text message inside the template so styles apply correctly
            ctx["message_html"] = default_message or ""
            balance_html = render_to_string(
                "programs/balance_sheet_email.html", ctx, request=None
            )
            full_html = balance_html
            try:
                inlined_html = transform(full_html)
            except Exception:
                inlined_html = full_html
            text_body = strip_tags(inlined_html)

            # Ensure dest is a list of flat email strings
            if isinstance(dest, str):
                to_list = [dest]
            else:
                to_list = list(dest)
            # Normalize: strip and drop empties/None
            to_list = [str(e).strip() for e in to_list if e and str(e).strip()]
            if not to_list:
                logger.warning(
                    "ProgramEmailBalances: no valid recipient emails for %s; skipping",
                    data["student"],
                )
                continue

            # Place all adult emails in To; only archive address in BCC
            to_addr = to_list
            bcc = ["swithee@andrew.cmu.edu"]

            email = EmailMultiAlternatives(
                subject=subject,
                body=text_body,
                from_email=from_email,
                to=to_addr,
                bcc=bcc,
                connection=connection,
            )
            email.attach_alternative(inlined_html, "text/html")
            try:
                sent = email.send(fail_silently=False)
                sent_total += sent
            except Exception as e:
                logger.error(
                    "ProgramEmailBalances: send failed for %s | error=%s",
                    data["student"],
                    e,
                    exc_info=True,
                )

        if test_email:
            messages.success(request, f"Test email sent to {test_email}.")
        else:
            messages.success(
                request, f"Balance emails queued/sent for {len(to_send)} student(s)."
            )
        return redirect("program_dues_owed", pk=program.pk)

    def _render(self, request, form, program):
        from django.shortcuts import render

        return render(
            request,
            self.template_name,
            {
                "program": program,
                "form": form,
            },
        )


class ProgramDuesOwedView(LoginRequiredMixin, LeadMentorRequiredMixin, View):
    """
    Lists all students enrolled in a specific program and the total amount each currently owes
    for that program, using the same balance computation as the per-program balance sheet.
    """

    template_name = "programs/dues_owed.html"
    section = "programs"

    def _program_balance_for_student(self, student, program):
        from .permission_views import can_user_read

        can_view_sliding = can_user_read(self.request.user, "sliding_scale")
        return get_student_program_balance(
            student,
            program,
            can_view_sliding=can_view_sliding,
        )

    def get(self, request, pk):
        from django.shortcuts import render

        program = get_object_or_404(Program, pk=pk)
        # Fetch all enrollments for this program.
        enrollments = (
            Enrollment.objects.filter(program=program)
            .select_related("student", "student__school")
            .order_by(
                Lower(
                    Coalesce(
                        NullIf("student__first_name", Value("")),
                        "student__legal_first_name",
                    )
                ),
                Lower("student__last_name"),
            )
        )

        active_rows = []
        inactive_rows = []
        grand_total = 0
        filter_owed = request.GET.get("filter") == "owed"
        for e in enrollments:
            s = e.student
            balance_sum = self._program_balance_for_student(s, program)
            if filter_owed and balance_sum <= 0:
                continue

            row = {
                "student": s,
                "amount_owed": balance_sum,
            }

            # A student is inactive if their enrollment is marked inactive,
            # or if the student record itself is marked graduated.
            if not e.active or s.graduated:
                inactive_rows.append(row)
            else:
                active_rows.append(row)

            grand_total += balance_sum

        return render(
            request,
            self.template_name,
            {
                "program": program,
                "active_rows": active_rows,
                "inactive_rows": inactive_rows,
                "rows": active_rows + inactive_rows,
                "grand_total": grand_total,
                "filter_owed": filter_owed,
            },
        )


class ProgramSignoutSheetView(LoginRequiredMixin, DynamicReadPermissionMixin, View):
    template_name = "programs/signout_sheet.html"
    section = "programs"

    def get(self, request, pk):
        from django.shortcuts import render

        program = get_object_or_404(Program, pk=pk)
        # Fetch students enrolled in the program, active first, then inactive
        base_qs = (
            program.students.select_related("user")
            .all()
            .annotate(
                sort_first=Lower(
                    Coalesce(NullIf("first_name", Value("")), "legal_first_name")
                ),
                sort_last=Lower("last_name"),
            )
        )
        students = list(
            base_qs.filter(
                enrollment__program=program, enrollment__active=True, graduated=False
            )
            .distinct()
            .order_by("sort_first", "sort_last")
        )
        ctx = {
            "program": program,
            "students": students,
        }
        return render(request, self.template_name, ctx)


class ProgramSchoolsView(LoginRequiredMixin, DynamicReadPermissionMixin, View):
    template_name = "programs/schools.html"
    section = "programs"

    def get(self, request, pk):
        from django.shortcuts import render

        program = get_object_or_404(Program, pk=pk)
        # Active (non-graduated) students enrolled in this program, grouped by school
        students = (
            Student.objects.filter(
                enrollment__program=program, enrollment__active=True, graduated=False
            )
            .distinct()
            .select_related("school")
            .annotate(
                sort_first=Coalesce(
                    NullIf("first_name", Value("")), "legal_first_name"
                ),
            )
            .order_by("school__name", Lower("sort_first"), Lower("last_name"))
        )
        grouped = {}
        for s in students:
            label = s.school.name if s.school_id else "No School"
            grouped.setdefault(label, []).append(s)
        grouped_items = sorted(
            grouped.items(), key=lambda kv: (kv[0] == "No School", kv[0] or "")
        )
        return render(
            request,
            self.template_name,
            {
                "program": program,
                "grouped": grouped_items,
            },
        )


class ProgramStudentMapView(LoginRequiredMixin, DynamicReadPermissionMixin, View):
    template_name = "programs/map.html"
    section = "programs"

    def get(self, request, pk):
        from django.shortcuts import render

        program = get_object_or_404(Program, pk=pk)
        # Active (non-graduated) students enrolled in this program with some address info
        students = (
            Student.objects.filter(
                enrollment__program=program, enrollment__active=True, graduated=False
            )
            .distinct()
            .only(
                "first_name",
                "legal_first_name",
                "last_name",
                "address",
                "city",
                "state",
                "zip_code",
            )
            .annotate(
                sort_first=Coalesce(NullIf("first_name", Value("")), "legal_first_name")
            )
            .order_by(Lower("sort_first"), Lower("last_name"))
        )
        items = []
        for s in students:
            parts = [s.address or "", s.city or "", s.state or "", s.zip_code or ""]
            addr = ", ".join([p for p in parts if p]).strip(", ")
            if not addr:
                continue
            name = f"{(s.first_name or s.legal_first_name or '').strip()} {s.last_name}".strip()
            items.append(
                {
                    "name": name or f"Student #{s.pk}",
                    "address": addr,
                }
            )
        return render(
            request,
            self.template_name,
            {
                "program": program,
                "items": items,
            },
        )


class AdultsListView(
    LoginRequiredMixin, DynamicReadPermissionMixin, SortableListViewMixin, ListView
):
    model = Adult
    template_name = "adults/list.html"
    context_object_name = "adults"
    section = "adult_info"

    sort_fields = {
        "name": (Lower("first_name"), Lower("last_name")),
        "email": Lower("personal_email"),
        "phone": "phone_number",
    }
    default_sort_field = "name"

    def get_queryset(self):

        qs = Adult.objects.all().prefetch_related("students")
        program_id = self.kwargs.get("program_id")
        if program_id:
            qs = qs.filter(students__enrollment__program_id=program_id).distinct()

        role = get_user_role(self.request.user)
        if role == "Parent":
            try:
                adult = self.request.user.adult_profile
                qs = qs.filter(pk=adult.pk)
            except (Adult.DoesNotExist, AttributeError):
                qs = Adult.objects.none()
        elif role == "Mentor":
            qs = qs.filter(
                is_parent=True, students__enrollment__program__active=True
            ).distinct()
        return self.apply_sorting(qs)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)

        ctx["role"] = get_user_role(self.request.user)
        program_id = self.kwargs.get("program_id")
        if program_id:
            ctx["program"] = get_object_or_404(Program, pk=program_id)
        return ctx


class AdultCreateView(
    PassUserToFormMixin,
    LogFormSaveMixin,
    LoginRequiredMixin,
    PermissionRequiredMixin,
    CreateView,
):
    model = Adult
    form_class = AdultForm
    template_name = "adults/form.html"
    permission_required = "programs.add_adult"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["back_url"] = self.request.META.get("HTTP_REFERER", "/")
        return ctx

    def get_success_url(self):
        return reverse("adult_list")


class AdultUpdateView(
    PassUserToFormMixin,
    SensitiveDataViewMixin,
    LogFormSaveMixin,
    LoginRequiredMixin,
    DynamicWritePermissionMixin,
    UpdateView,
):
    model = Adult
    form_class = AdultForm
    template_name = "adults/form.html"
    permission_required = "programs.change_adult"
    section = "adult_info"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["back_url"] = self.request.META.get("HTTP_REFERER", "/")
        return ctx

    def get_success_url(self):
        nxt = self.request.GET.get("next")
        safe_url = get_safe_url(self.request, nxt)
        if safe_url:
            return safe_url
        return reverse("adult_list")


# --- Program documents (Step 9 blank forms) ---------------------------------


class ProgramDocumentCreateView(
    LogFormSaveMixin,
    LoginRequiredMixin,
    PermissionRequiredMixin,
    CreateView,
):
    """Add a blank document (e.g. PDF) to a Program. Shown on the Program
    settings page so lead mentors can manage them without going through
    the Django admin.
    """

    permission_required = "programs.change_program"
    model = ProgramDocument
    form_class = ProgramDocumentForm
    template_name = "programs/program_document_form.html"

    def dispatch(self, request, *args, **kwargs):
        self.program = get_object_or_404(Program, pk=kwargs["pk"])
        return super().dispatch(request, *args, **kwargs)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["program"] = self.program
        return kwargs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["program"] = self.program
        ctx["is_create"] = True
        return ctx

    def form_valid(self, form):
        messages.success(self.request, "Document added.")
        return super().form_valid(form)

    def get_success_url(self):
        return reverse("program_detail", args=[self.program.pk])


class ProgramDocumentUpdateView(
    LogFormSaveMixin,
    LoginRequiredMixin,
    PermissionRequiredMixin,
    UpdateView,
):
    permission_required = "programs.change_program"
    model = ProgramDocument
    form_class = ProgramDocumentForm
    template_name = "programs/program_document_form.html"
    pk_url_kwarg = "doc_id"

    def dispatch(self, request, *args, **kwargs):
        self.program = get_object_or_404(Program, pk=kwargs["pk"])
        return super().dispatch(request, *args, **kwargs)

    def get_object(self, queryset=None):
        return get_object_or_404(
            ProgramDocument, pk=self.kwargs["doc_id"], program=self.program
        )

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["program"] = self.program
        return kwargs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["program"] = self.program
        ctx["is_create"] = False
        return ctx

    def form_valid(self, form):
        messages.success(self.request, "Document updated.")
        return super().form_valid(form)

    def get_success_url(self):
        return reverse("program_detail", args=[self.program.pk])


class ProgramDocumentDeleteView(
    LoginRequiredMixin,
    PermissionRequiredMixin,
    View,
):
    """Delete a Program Document. POST-only (with a JS confirm on the
    detail page); GET renders a small confirmation page for safety.
    """

    permission_required = "programs.change_program"
    template_name = "programs/program_document_confirm_delete.html"

    def dispatch(self, request, *args, **kwargs):
        self.program = get_object_or_404(Program, pk=kwargs["pk"])
        self.document = get_object_or_404(
            ProgramDocument, pk=kwargs["doc_id"], program=self.program
        )
        return super().dispatch(request, *args, **kwargs)

    def get(self, request, *args, **kwargs):
        from django.shortcuts import render

        return render(
            request,
            self.template_name,
            {"program": self.program, "document": self.document},
        )

    def post(self, request, *args, **kwargs):
        name = self.document.name
        self.document.delete()
        messages.success(request, f"Deleted document “{name}”.")
        return redirect("program_detail", pk=self.program.pk)
