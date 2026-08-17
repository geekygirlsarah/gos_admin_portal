from io import BytesIO

from django.contrib.auth.models import Group, Permission, User
from django.test import TestCase
from django.urls import reverse
from openpyxl import load_workbook

from programs.models import Enrollment, Program, Student


class ProgramStudentExportViewTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(
            username="leadmentor", password="pass12345"  # nosec B106
        )
        group, _ = Group.objects.get_or_create(name="LeadMentor")
        self.user.groups.add(group)
        self.perm = Permission.objects.get(codename="view_student")
        self.user.user_permissions.add(self.perm)
        self.client.login(username="leadmentor", password="pass12345")  # nosec B106

        self.program = Program.objects.create(name="Robotics", active=True)

        self.student_a = Student.objects.create(
            legal_first_name="Alice", last_name="Smith", graduation_year=2029
        )
        self.student_b = Student.objects.create(
            legal_first_name="Bob", last_name="Jones", graduation_year=2031
        )
        self.student_c = Student.objects.create(
            legal_first_name="Charlie", last_name="Adams"
        )

        Enrollment.objects.create(
            student=self.student_a, program=self.program, active=True
        )
        Enrollment.objects.create(
            student=self.student_b, program=self.program, active=True
        )
        Enrollment.objects.create(
            student=self.student_c, program=self.program, active=False
        )

    def test_export_requires_login(self):
        self.client.logout()
        url = reverse("program_student_export", args=[self.program.pk])
        resp = self.client.get(url)
        self.assertEqual(resp.status_code, 302)
        self.assertIn("/accounts/login/", resp.url)

    def test_export_returns_xlsx(self):
        url = reverse("program_student_export", args=[self.program.pk])
        resp = self.client.get(url)
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(
            resp["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn(".xlsx", resp["Content-Disposition"])

    def test_export_includes_only_active_enrolled_students(self):
        url = reverse("program_student_export", args=[self.program.pk])
        resp = self.client.get(url)
        wb = load_workbook(BytesIO(resp.content))
        ws = wb.active

        headers = [cell.value for cell in ws[1]]
        self.assertEqual(headers, ["First Name", "Last Name", "Grade"])

        names = [(row[0].value, row[1].value) for row in ws.iter_rows(min_row=2)]
        first_names = {n[0] for n in names}
        self.assertIn("Alice", first_names)
        self.assertIn("Bob", first_names)
        self.assertNotIn("Charlie", first_names)

    def test_export_includes_grade(self):
        url = reverse("program_student_export", args=[self.program.pk])
        resp = self.client.get(url)
        wb = load_workbook(BytesIO(resp.content))
        ws = wb.active

        rows = {row[0].value: row[2].value for row in ws.iter_rows(min_row=2)}
        # Alice: grad 2029 → should have a grade
        self.assertIsNotNone(rows["Alice"])
        self.assertNotEqual(rows["Alice"], "")
        # Charlie: no graduation_year → grade is empty string
        self.assertNotIn("Charlie", rows)

    def test_export_students_sorted_alphabetically(self):
        url = reverse("program_student_export", args=[self.program.pk])
        resp = self.client.get(url)
        wb = load_workbook(BytesIO(resp.content))
        ws = wb.active

        first_names = [row[0].value for row in ws.iter_rows(min_row=2)]
        self.assertEqual(first_names, sorted(first_names))

    def test_export_filename_contains_program_name(self):
        url = reverse("program_student_export", args=[self.program.pk])
        resp = self.client.get(url)
        self.assertIn("Robotics", resp["Content-Disposition"])

    def test_404_for_nonexistent_program(self):
        url = reverse("program_student_export", args=[9999])
        resp = self.client.get(url)
        self.assertIn(resp.status_code, (302, 404))
