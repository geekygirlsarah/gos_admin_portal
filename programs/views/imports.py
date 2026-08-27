from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.shortcuts import redirect, render
from django.views.generic import View

from ..constants import RELATIONSHIP_CHOICES
from ..models import (
    Adult,
    AdultStudentRelationship,
    Program,
    RaceEthnicity,
    School,
    SchoolDistrict,
    Student,
)
from ..utils import redirect_back
from .mixins import logger


class ImportDashboardView(LoginRequiredMixin, View):
    def get(self, request):
        programs = Program.objects.all().order_by("name")
        programs_with_attendance = [p for p in programs if p.has_feature("attendance")]
        return render(
            request,
            "imports/dashboard.html",
            {
                "programs": programs,
                "attendance_programs": programs_with_attendance,
            },
        )


class StudentImportView(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = "programs.add_student"

    def post(self, request):
        file = request.FILES.get("file")
        if not file:
            messages.error(request, "No file uploaded.")
            return redirect_back(request, "import_dashboard")
        name = file.name.lower()
        overwrite = request.POST.get("overwrite") == "1"
        created = 0
        updated = 0
        errors = 0
        try:
            if name.endswith(".csv"):
                import csv
                import io

                text = io.TextIOWrapper(file.file, encoding="utf-8")
                reader = csv.DictReader(text)
                rows = list(reader)
            elif name.endswith(".xlsx"):
                from openpyxl import load_workbook

                wb = load_workbook(filename=file, read_only=True, data_only=True)
                ws = wb.active
                headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
                rows = []
                for r in ws.iter_rows(min_row=2, values_only=True):
                    rows.append(
                        {
                            str(headers[i]): r[i]
                            for i in range(len(headers))
                            if headers[i] is not None
                        }
                    )
            else:
                messages.error(
                    request, "Unsupported file type. Please upload CSV or XLSX."
                )
                return redirect("import_dashboard")

            # Helpers
            from datetime import date, datetime

            def raw(d, *keys):
                for k in keys:
                    if k in d and d[k] is not None:
                        return d[k]
                return None

            def val(d, *keys):
                for k in keys:
                    if k in d and d[k] is not None:
                        v = str(d[k]).strip()
                        if v != "" and v.lower() != "none":
                            return v
                return None

            def val_bool(d, *keys):
                v = val(d, *keys)
                if v is None:
                    return None
                s = v.strip().lower()
                if s in ("y", "yes", "true", "t", "1"):
                    return True
                if s in ("n", "no", "false", "f", "0"):
                    return False
                return None

            def val_date(d, *keys):
                # Accept date objects from XLSX or parse common string formats
                rv = raw(d, *keys)
                if isinstance(rv, datetime):
                    return rv.date()
                if isinstance(rv, date):
                    return rv
                v = val(d, *keys)
                if not v:
                    return None
                for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
                    try:
                        return datetime.strptime(v, fmt).date()
                    except ValueError:
                        continue
                return None

            def get_or_create_parent(first, last, email):
                # Try to find by email first
                if email:
                    p = Adult.objects.filter(personal_email__iexact=email).first()
                    if p:
                        if overwrite:
                            changed_parent = False
                            if first and p.legal_first_name != first:
                                p.legal_first_name = first
                                changed_parent = True
                            if last and p.last_name != last:
                                p.last_name = last
                                changed_parent = True
                            if changed_parent:
                                p.save()
                        return p
                # Next try by name match
                if first and last:
                    p = Adult.objects.filter(
                        legal_first_name__iexact=first, last_name__iexact=last
                    ).first()
                    if p:
                        if (
                            overwrite
                            and email
                            and (p.personal_email or "").lower()
                            != (email or "").lower()
                        ):
                            p.personal_email = email
                            p.save()
                        return p
                # If we have at least one of name or email, create
                if first or last or email:
                    return Adult.objects.create(
                        legal_first_name=first
                        or (email.split("@")[0] if email else "Parent"),
                        last_name=last or "(contact)",
                        personal_email=email or None,
                        is_parent=True,
                    )
                return None

            for d in rows:
                first = val(d, "first_name", "First Name", "Preferred First Name")
                legal_first = val(d, "legal_first_name", "Legal First Name") or first
                last = val(d, "last_name", "Last Name")
                if not last or not legal_first:
                    errors += 1
                    continue

                # Simple strings
                pronouns = val(d, "pronouns", "Pronouns")
                address = val(d, "address", "Address", "Street Address")
                city = val(d, "city", "City")
                state = val(d, "state", "State")
                zip_code = val(d, "zip_code", "Zip Code", "ZIP", "Zip")
                cell_phone = val(
                    d,
                    "cell_phone_number",
                    "Cell Phone Number",
                    "Cell Phone",
                    "Phone",
                    "Phone Number",
                )
                personal_email = val(d, "personal_email", "Email", "Personal Email")
                andrew_id = val(d, "andrew_id", "Andrew ID", "AndrewID")
                andrew_email = val(d, " andrew_email", "Andrew Email")
                race_ethnicity = val(
                    d, "race_ethnicity", "Race/Ethnicity", "Race", "Ethnicity"
                )
                tshirt_size = val(d, "tshirt_size", "T-Shirt Size", "Shirt Size")
                discord_handle = val(
                    d, "discord_handle", "Discord Handle", "Discord", "Discord Username"
                )

                # Dates and booleans
                dob = val_date(d, "date_of_birth", "Date of Birth", "DOB", "Birthdate")
                seen_once = val_bool(d, "seen_once", "Seen Once")
                on_discord = val_bool(d, "on_discord", "On Discord")
                graduated = val_bool(d, "graduated", "Graduated")
                # Backward compatibility for older templates that still send Active.
                # Student now uses "graduated" instead of an "active" field.
                active = val_bool(d, "active", "Active")
                if graduated is None and active is not None:
                    graduated = not active

                # School/year
                school_name = val(d, "school", "School")
                grad = val(d, "graduation_year", "Graduation Year")
                school = None
                if school_name:
                    school, _ = School.objects.get_or_create(name=school_name)
                grad_year = None
                if grad and str(grad).isdigit():
                    grad_year = int(str(grad))

                obj, created_flag = Student.objects.get_or_create(
                    last_name=last,
                    legal_first_name=legal_first,
                    defaults={
                        "preferred_first_name": first if first != legal_first else None,
                        "pronouns": pronouns,
                        "date_of_birth": dob,
                        "address": address,
                        "city": city,
                        "state": state,
                        "zip_code": zip_code,
                        "phone_number": cell_phone,
                        "phone_type": "cell",
                        "can_receive_texts": True,
                        "personal_email": personal_email,
                        "andrew_id": andrew_id,
                        "andrew_email": andrew_email,
                        "tshirt_size": tshirt_size,
                        "seen_once": seen_once if seen_once is not None else False,
                        "on_discord": on_discord if on_discord is not None else False,
                        "discord_handle": discord_handle,
                        "school": school,
                        "graduation_year": grad_year,
                        "graduated": graduated if graduated is not None else False,
                    },
                )
                if created_flag:
                    created += 1
                elif overwrite:
                    changed = False
                    # Strings and relations
                    for field, value in [
                        ("preferred_first_name", first),
                        ("pronouns", pronouns),
                        ("address", address),
                        ("city", city),
                        ("state", state),
                        ("zip_code", zip_code),
                        ("phone_number", cell_phone),
                        ("phone_type", "cell"),
                        ("can_receive_texts", True),
                        ("personal_email", personal_email),
                        ("andrew_id", andrew_id),
                        ("andrew_email", andrew_email),
                        ("tshirt_size", tshirt_size),
                        ("discord_handle", discord_handle),
                    ]:
                        if value and getattr(obj, field) != value:
                            setattr(obj, field, value)
                            changed = True
                    if dob and obj.date_of_birth != dob:
                        obj.date_of_birth = dob
                        changed = True
                    if school and obj.school != school:
                        obj.school = school
                        changed = True
                    if grad_year and obj.graduation_year != grad_year:
                        obj.graduation_year = grad_year
                        changed = True
                    # Booleans (allow False updates)
                    if seen_once is not None and obj.seen_once != seen_once:
                        obj.seen_once = seen_once
                        changed = True
                    if on_discord is not None and obj.on_discord != on_discord:
                        obj.on_discord = on_discord
                        changed = True
                    if graduated is not None and obj.graduated != graduated:
                        obj.graduated = graduated
                        changed = True
                    if changed:
                        obj.save()
                        updated += 1

                # Map race/ethnicity text to multi-select options
                try:
                    opts = RaceEthnicity.match_from_text(race_ethnicity)
                    if opts.exists():
                        obj.race_ethnicities.set(list(opts))
                except Exception:
                    logger.debug(
                        "Race/Ethnicity matching failed during import", exc_info=True
                    )

                # Parent linkage (primary and secondary)
                prim_first = val(
                    d,
                    "primary_parent_first_name",
                    "Primary Parent First Name",
                    "Primary First Name",
                    "Primary First",
                )
                prim_last = val(
                    d,
                    "primary_parent_last_name",
                    "Primary Parent Last Name",
                    "Primary Last Name",
                    "Primary Last",
                )
                prim_email = val(
                    d,
                    "primary_parent_email",
                    "Primary Parent Email",
                    "Primary Email",
                    "Primary E-mail",
                    "Primary Email Address",
                )
                sec_first = val(
                    d,
                    "secondary_parent_first_name",
                    "Secondary Parent First Name",
                    "Secondary First Name",
                    "Secondary First",
                )
                sec_last = val(
                    d,
                    "secondary_parent_last_name",
                    "Secondary Parent Last Name",
                    "Secondary Last Name",
                    "Secondary Last",
                )
                sec_email = val(
                    d,
                    "secondary_parent_email",
                    "Secondary Parent Email",
                    "Secondary Email",
                    "Secondary E-mail",
                    "Secondary Email Address",
                )

                contact_changed = False
                primary = get_or_create_parent(prim_first, prim_last, prim_email)
                secondary = get_or_create_parent(sec_first, sec_last, sec_email)
                if primary:
                    if obj.primary_contact_id != getattr(primary, "id", None):
                        obj.primary_contact = primary
                        contact_changed = True
                    # Ensure M2M link exists (both sides)
                    if primary.id and not obj.adults.filter(id=primary.id).exists():
                        obj.adults.add(primary)
                        primary.students.add(obj)
                if secondary:
                    if obj.secondary_contact_id != getattr(secondary, "id", None):
                        obj.secondary_contact = secondary
                        contact_changed = True
                    if secondary.id and not obj.adults.filter(id=secondary.id).exists():
                        obj.adults.add(secondary)
                        secondary.students.add(obj)
                if contact_changed:
                    obj.save(
                        update_fields=[
                            "primary_contact_relationship",
                            "secondary_contact_relationship",
                            "updated_at",
                        ]
                    )
                    if not created_flag:
                        # Only count as updated when not newly created and not already counted
                        updated += 1
            if created or updated:
                messages.success(
                    request,
                    f"Imported {created} new, updated {updated}. Skipped {errors}.",
                )
            else:
                messages.info(request, "No rows imported.")
        except Exception as e:
            messages.error(request, f"Import failed: {e}")
        return redirect_back(request, "import_dashboard")


class ParentImportView(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = "programs.add_adult"

    def post(self, request):
        file = request.FILES.get("file")
        if not file:
            messages.error(request, "No file uploaded.")
            return redirect_back(request, "import_dashboard")
        name = file.name.lower()
        overwrite = request.POST.get("overwrite") == "1"
        created = 0
        updated = 0
        errors = 0
        try:
            if name.endswith(".csv"):
                import csv
                import io

                text = io.TextIOWrapper(file.file, encoding="utf-8")
                reader = csv.DictReader(text)
                rows = list(reader)
            elif name.endswith(".xlsx"):
                from openpyxl import load_workbook

                wb = load_workbook(filename=file, read_only=True)
                ws = wb.active
                headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
                rows = []
                for r in ws.iter_rows(min_row=2, values_only=True):
                    rows.append(
                        {
                            str(headers[i]): r[i]
                            for i in range(len(headers))
                            if headers[i] is not None
                        }
                    )
            else:
                messages.error(
                    request, "Unsupported file type. Please upload CSV or XLSX."
                )
                return redirect("import_dashboard")

            def val(d, *keys):
                for k in keys:
                    if k in d and d[k] is not None:
                        v = str(d[k]).strip()
                        if v != "" and v.lower() != "none":
                            return v
                return None

            for d in rows:
                first = val(d, "first_name", "First Name")
                last = val(d, "last_name", "Last Name")
                if not first or not last:
                    errors += 1
                    continue
                email = val(d, "email", "Email")
                phone = val(
                    d,
                    "cell_phone",
                    "Cell Phone",
                    "Cell Phone Number",
                    "Phone",
                    "Phone Number",
                )
                obj, created_flag = Adult.objects.get_or_create(
                    legal_first_name=first,
                    last_name=last,
                    defaults={
                        "personal_email": email,
                        "phone_number": phone,
                        "phone_type": "cell",
                        "can_receive_texts": True,
                        "is_parent": True,
                    },
                )
                if created_flag:
                    created += 1
                else:
                    changed = False
                    if not obj.is_parent:
                        obj.is_parent = True
                        changed = True

                    if not overwrite:
                        if changed:
                            obj.save(update_fields=["is_parent", "updated_at"])
                            updated += 1
                        continue

                    if email and obj.personal_email != email:
                        obj.personal_email = email
                        changed = True
                    if phone and obj.phone_number != phone:
                        obj.phone_number = phone
                        obj.phone_type = "cell"
                        obj.can_receive_texts = True
                        changed = True
                    if changed:
                        obj.save()
                        updated += 1
            messages.success(
                request, f"Imported {created} new, updated {updated}. Skipped {errors}."
            )
        except Exception as e:
            messages.error(request, f"Import failed: {e}")
        return redirect_back(request, "import_dashboard")


class RelationshipImportView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """
    Re-link existing Students to Parent/Adult records and set relationship types.
    Safe to run multiple times (idempotent). Optionally supports dry-run.
    """

    permission_required = "programs.change_student"

    def post(self, request):
        file = request.FILES.get("file")
        if not file:
            messages.error(request, "No file uploaded.")
            return redirect_back(request, "import_dashboard")
        name = file.name.lower()
        dry_run = request.POST.get("dry_run") in ("1", "on", "true", "True")
        overwrite = request.POST.get("overwrite") == "1"
        can_create_parents = request.user.has_perm("programs.add_adult")

        linked = 0
        set_primary = 0
        set_secondary = 0
        rel_updated = 0
        created_parents = 0
        would_create_parents = 0
        missing_or_ambiguous_students = 0
        skipped = 0

        try:
            # Parse CSV/XLSX similar to other imports
            if name.endswith(".csv"):
                import csv
                import io

                text = io.TextIOWrapper(file.file, encoding="utf-8")
                reader = csv.DictReader(text)
                rows = list(reader)
            elif name.endswith(".xlsx"):
                from openpyxl import load_workbook

                wb = load_workbook(filename=file, read_only=True, data_only=True)
                ws = wb.active
                headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
                rows = []
                for r in ws.iter_rows(min_row=2, values_only=True):
                    rows.append(
                        {
                            str(headers[i]): r[i]
                            for i in range(len(headers))
                            if headers[i] is not None
                        }
                    )
            else:
                messages.error(
                    request, "Unsupported file type. Please upload CSV or XLSX."
                )
                return redirect("import_dashboard")

            # Helpers
            from datetime import date, datetime

            def raw(d, *keys):
                for k in keys:
                    if k in d and d[k] is not None:
                        return d[k]
                return None

            def val(d, *keys):
                for k in keys:
                    if k in d and d[k] is not None:
                        v = str(d[k]).strip()
                        if v != "" and v.lower() != "none":
                            return v
                return None

            def val_date(d, *keys):
                rv = raw(d, *keys)
                if isinstance(rv, datetime):
                    return rv.date()
                if isinstance(rv, date):
                    return rv
                s = val(d, *keys)
                if not s:
                    return None
                for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
                    try:
                        return datetime.strptime(s, fmt).date()
                    except ValueError:
                        continue
                return None

            def normalize_rel(s):
                if not s:
                    return None
                s2 = s.strip().lower()
                # Accept either key or display label
                keys = {k for k, _ in RELATIONSHIP_CHOICES}
                labels = {lbl.lower(): k for k, lbl in RELATIONSHIP_CHOICES}
                synonyms = {
                    "mom": "mother",
                    "dad": "father",
                    "grandma": "grandmother",
                    "grandpa": "grandfather",
                    "guardian": "guardian",
                    "parent": "parent",
                }
                if s2 in keys:
                    return s2
                if s2 in labels:
                    return labels[s2]
                if s2 in synonyms and synonyms[s2] in keys:
                    return synonyms[s2]
                return None

            def resolve_student(d):
                # Priority: ID -> Andrew ID -> (First/Legal First + Last + DOB) -> (First/Legal First + Last)
                sid = val(d, "student_id", "Student ID", "ID")
                if sid and str(sid).isdigit():
                    st = Student.objects.filter(pk=int(str(sid))).first()
                    if st:
                        return st

                aid = val(d, "andrew_id", "Andrew ID", "AndrewID")
                if aid:
                    st = Student.objects.filter(andrew_id__iexact=aid).first()
                    if st:
                        return st

                last = val(d, "last_name", "Last Name")
                first = val(d, "first_name", "First Name", "Preferred First Name")
                legal_first = val(d, "legal_first_name", "Legal First Name") or first
                dob = val_date(d, "date_of_birth", "Date of Birth", "DOB", "Birthdate")

                if not last or not legal_first:
                    return None

                qs = Student.objects.filter(
                    last_name__iexact=last, legal_first_name__iexact=legal_first
                )
                if dob:
                    qs = qs.filter(date_of_birth=dob)
                count = qs.count()
                if count == 1:
                    return qs.first()
                if count == 0 and first and first != legal_first:
                    # Try match on preferred first + last (+dob)
                    qs = Student.objects.filter(
                        last_name__iexact=last, preferred_first_name__iexact=first
                    )
                    if dob:
                        qs = qs.filter(date_of_birth=dob)
                    if qs.count() == 1:
                        return qs.first()
                return None if qs.count() != 1 else qs.first()

            def find_or_create_parent(first, last, email):
                # Try resolve by email first
                p = None
                if email:
                    p = Adult.objects.filter(personal_email__iexact=email).first()
                if not p and first and last:
                    p = Adult.objects.filter(
                        legal_first_name__iexact=first, last_name__iexact=last
                    ).first()
                created = False
                if not p:
                    if dry_run or not can_create_parents:
                        return None, False, True  # would create
                    p = Adult.objects.create(
                        legal_first_name=first
                        or (email.split("@")[0] if email else "Parent"),
                        last_name=last or "(contact)",
                        personal_email=email or None,
                        is_parent=True,
                    )
                    created = True
                else:
                    # If we found existing Adult but not flagged as parent, set it
                    if not dry_run and not p.is_parent:
                        p.is_parent = True
                        p.save(update_fields=["is_parent", "updated_at"])
                return p, created, False

            for d in rows:
                student = resolve_student(d)
                if not student:
                    missing_or_ambiguous_students += 1
                    continue

                groups = [
                    {
                        "role": "primary",
                        "first": val(
                            d,
                            "primary_parent_first_name",
                            "Primary Parent First Name",
                            "Primary First Name",
                            "Primary First",
                        ),
                        "last": val(
                            d,
                            "primary_parent_last_name",
                            "Primary Parent Last Name",
                            "Primary Last Name",
                            "Primary Last",
                        ),
                        "email": val(
                            d,
                            "primary_parent_email",
                            "Primary Parent Email",
                            "Primary Email",
                        ),
                        "rel": val(
                            d,
                            "primary_parent_relationship",
                            "Primary Parent Relationship",
                            "Primary Relationship",
                        ),
                    },
                    {
                        "role": "secondary",
                        "first": val(
                            d,
                            "secondary_parent_first_name",
                            "Secondary Parent First Name",
                            "Secondary First Name",
                            "Secondary First",
                        ),
                        "last": val(
                            d,
                            "secondary_parent_last_name",
                            "Secondary Parent Last Name",
                            "Secondary Last Name",
                            "Secondary Last",
                        ),
                        "email": val(
                            d,
                            "secondary_parent_email",
                            "Secondary Parent Email",
                            "Secondary Email",
                        ),
                        "rel": val(
                            d,
                            "secondary_parent_relationship",
                            "Secondary Parent Relationship",
                            "Secondary Relationship",
                        ),
                    },
                ]

                updated_student_fields = set()

                for g in groups:
                    if not (g["first"] or g["last"] or g["email"]):
                        continue
                    adult, created_flag, would_create = find_or_create_parent(
                        g["first"], g["last"], g["email"]
                    )
                    if would_create:
                        would_create_parents += 1
                        continue
                    if created_flag:
                        created_parents += 1
                    if not adult:
                        skipped += 1
                        continue

                    # Relationship type
                    rel_key = normalize_rel(g["rel"])
                    rel_key = rel_key or "parent"
                    if not dry_run:
                        _, rel_created = (
                            AdultStudentRelationship.objects.update_or_create(
                                adult=adult,
                                student=student,
                                defaults={"relationship_to_student": rel_key},
                            )
                        )
                        if rel_created:
                            linked += 1
                    else:
                        if not student.adults.filter(id=adult.id).exists():
                            linked += 1
                    rel_updated += 1

                    # Ensure Adult is linked to Student (M2M) - handled by update_or_create when not dry_run.

                    # Optionally set primary/secondary contact
                    if g["role"] == "primary":
                        if student.primary_contact_id != adult.id:
                            if not dry_run and overwrite:
                                student.primary_contact = adult
                                updated_student_fields.add(
                                    "primary_contact_relationship"
                                )
                            set_primary += 1
                    elif g["role"] == "secondary":
                        if student.secondary_contact_id != adult.id:
                            if not dry_run and overwrite:
                                student.secondary_contact = adult
                                updated_student_fields.add(
                                    "secondary_contact_relationship"
                                )
                            set_secondary += 1

                if updated_student_fields and not dry_run:
                    fields = list(updated_student_fields) + ["updated_at"]
                    student.save(update_fields=fields)

            # Compose message
            notes = []
            if dry_run:
                notes.append("DRY RUN (no changes saved)")
            if not can_create_parents:
                notes.append(
                    "Note: lacking permission to create parents; rows requiring new parent were skipped."
                )
            extras = f" {'; '.join(notes)}" if notes else ""
            messages.success(
                request,
                f"Relationships import: linked {linked} (primary set {set_primary}, secondary set {set_secondary}); "
                f"updated relationship types {rel_updated}; "
                f"created parents {created_parents}"
                f"{(' (would create: ' + str(would_create_parents) + ')' if dry_run else '')}; "
                f"missing/ambiguous students {missing_or_ambiguous_students}; skipped {skipped}.{extras}",
            )
        except Exception as e:
            messages.error(request, f"Import failed: {e}")
        return redirect_back(request, "import_dashboard")


class MentorImportView(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = "programs.add_adult"

    def post(self, request):
        file = request.FILES.get("file")
        if not file:
            messages.error(request, "No file uploaded.")
            return redirect_back(request, "import_dashboard")
        name = file.name.lower()
        overwrite = request.POST.get("overwrite") == "1"
        created = 0
        updated = 0
        errors = 0
        try:
            if name.endswith(".csv"):
                import csv
                import io

                text = io.TextIOWrapper(file.file, encoding="utf-8")
                reader = csv.DictReader(text)
                rows = list(reader)
            elif name.endswith(".xlsx"):
                from openpyxl import load_workbook

                wb = load_workbook(filename=file, read_only=True)
                ws = wb.active
                headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
                rows = []
                for r in ws.iter_rows(min_row=2, values_only=True):
                    rows.append(
                        {
                            str(headers[i]): r[i]
                            for i in range(len(headers))
                            if headers[i] is not None
                        }
                    )
            else:
                messages.error(
                    request, "Unsupported file type. Please upload CSV or XLSX."
                )
                return redirect("import_dashboard")

            def val(d, *keys):
                for k in keys:
                    if k in d and d[k] is not None:
                        v = str(d[k]).strip()
                        if v != "" and v.lower() != "none":
                            return v
                return None

            for d in rows:
                first = val(d, "first_name", "First Name")
                last = val(d, "last_name", "Last Name")
                if not first or not last:
                    errors += 1
                    continue
                email = val(d, "personal_email", "Email", "Personal Email")
                andrew_email = val(d, "andrew_email", "Andrew Email")
                role = val(d, "role", "Role") or "mentor"
                obj, created_flag = Adult.objects.get_or_create(
                    legal_first_name=first,
                    last_name=last,
                    defaults={
                        "personal_email": email,
                        "andrew_email": andrew_email,
                        "role": role,
                        "is_mentor": True,
                    },
                )
                if created_flag:
                    created += 1
                else:
                    changed = False
                    if not obj.is_mentor:
                        obj.is_mentor = True
                        changed = True

                    if not overwrite:
                        if changed:
                            obj.save(update_fields=["is_mentor", "updated_at"])
                            updated += 1
                        continue

                    for field, value in [
                        ("personal_email", email),
                        ("andrew_email", andrew_email),
                        ("role", role),
                    ]:
                        if value and getattr(obj, field) != value:
                            setattr(obj, field, value)
                            changed = True
                    if changed:
                        obj.save()
                        updated += 1
            messages.success(
                request, f"Imported {created} new, updated {updated}. Skipped {errors}."
            )
        except Exception as e:
            messages.error(request, f"Import failed: {e}")
        return redirect_back(request, "import_dashboard")


class SchoolImportView(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = "programs.add_school"

    def post(self, request):
        file = request.FILES.get("file")
        if not file:
            messages.error(request, "No file uploaded.")
            return redirect_back(request, "import_dashboard")
        name = file.name.lower()
        overwrite = request.POST.get("overwrite") == "1"
        created = 0
        updated = 0
        errors = 0
        try:
            if name.endswith(".csv"):
                import csv
                import io

                text = io.TextIOWrapper(file.file, encoding="utf-8")
                reader = csv.DictReader(text)
                rows = list(reader)
            elif name.endswith(".xlsx"):
                from openpyxl import load_workbook

                wb = load_workbook(filename=file, read_only=True, data_only=True)
                ws = wb.active
                headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
                rows = []
                for r in ws.iter_rows(min_row=2, values_only=True):
                    rows.append(
                        {
                            str(headers[i]): r[i]
                            for i in range(len(headers))
                            if headers[i] is not None
                        }
                    )
            else:
                messages.error(
                    request, "Unsupported file type. Please upload CSV or XLSX."
                )
                return redirect("import_dashboard")

            def val(d, *keys):
                for k in keys:
                    if k in d and d[k] is not None:
                        v = str(d[k]).strip()
                        if v != "" and v.lower() != "none":
                            return v
                return None

            for d in rows:
                school_name = val(d, "name", "Name", "School")
                if not school_name:
                    errors += 1
                    continue
                district = val(d, "district", "District", "School District")
                street = val(d, "street_address", "Street", "Street Address", "Address")
                city = val(d, "city", "City")
                state = val(d, "state", "State")
                zip_code = val(
                    d, "zip", "ZIP", "Zip", "zip_code", "Zip Code", "Postal Code"
                )
                obj, created_flag = School.objects.get_or_create(
                    name=school_name,
                    defaults={
                        "street_address": street,
                        "city": city,
                        "state": state,
                        "zip_code": zip_code,
                    },
                )
                if district:
                    district_obj, _ = SchoolDistrict.objects.get_or_create(
                        name=district
                    )
                    obj.district = district_obj
                if created_flag:
                    if district:
                        obj.save()
                    created += 1
                elif overwrite:
                    changed = False
                    for field, value in [
                        ("street_address", street),
                        ("city", city),
                        ("state", state),
                        ("zip_code", zip_code),
                    ]:
                        if value and getattr(obj, field) != value:
                            setattr(obj, field, value)
                            changed = True
                    if district and obj.district_id != district_obj.pk:
                        obj.district = district_obj
                        changed = True
                    if changed:
                        obj.save()
                        updated += 1
                elif district and obj.district_id != district_obj.pk:
                    obj.save()
                    updated += 1
            messages.success(
                request, f"Imported {created} new, updated {updated}. Skipped {errors}."
            )
        except Exception as e:
            messages.error(request, f"Import failed: {e}")
        return redirect_back(request, "import_dashboard")
